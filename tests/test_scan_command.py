"""La commande /scan : quels paris elle retient, et comment elle les rend.

La sélection est la seule partie qui porte de la logique, et elle est pure —
elle reçoit des lignes, un ensemble de marchés joués et une horloge. Le reste
(getUpdates, sendMessage) n'est que de la plomberie.
"""
from __future__ import annotations

from datetime import datetime, timedelta, timezone

import pytest

from bot_listener import format_scan, is_premium, select_playable
from src.alerter import TelegramConfig


NOW = datetime(2026, 8, 4, 12, 0, tzinfo=timezone.utc)


def _cfg(**kw) -> TelegramConfig:
    base = dict(bot_token="t", chat_id="c", premium_chat_id="prem",
                critical_chat_id="crit", min_minutes_to_kickoff=15)
    base.update(kw)
    return TelegramConfig(**base)


def _row(*, ev=12.0, odd=2.40, book="unibet_be", market="h2h", outcome="home",
         line=None, hours=3.0, home="Anderlecht", away="Club Brugge",
         sport="soccer"):
    """Une ligne value_bets telle que la requête la renvoie (sqlite3.Row-like)."""
    start = NOW + timedelta(hours=hours)
    return {
        "event_key": f"{start.strftime('%Y%m%d%H%M')}::"
                     f"{home.lower().replace(' ', '')}__vs__{away.lower().replace(' ', '')}",
        "book": book, "market": market, "outcome_label": outcome, "line": line,
        "odd_taken": odd, "ev_pct": ev, "kelly_pct": 1.5, "sport": sport,
    }


def _sel(rows, played=frozenset(), cfg=None):
    return select_playable(rows, set(played), cfg or _cfg(), NOW)


def _texts(bets, **kw):
    """Les textes seuls — format_scan renvoie (texte, paris du message)."""
    return [t for t, _ in format_scan(bets, now=NOW, **kw)]


# --------------------------------------------------------------- filtres ----

def test_played_market_is_excluded_including_its_other_outcomes():
    """Le cœur de la demande : ce qui a été joué ne réapparaît pas.

    Et pas seulement la sélection jouée — tout le marché, comme pour les
    alertes : avoir pris le 1 d'un 1X2 rend le X et le 2 sans objet."""
    home, draw = _row(outcome="home"), _row(outcome="draw")
    assert len(_sel([home, draw])) == 2          # rien de joué : les deux sortent
    played = {f"{home['event_key']}|h2h|None"}
    assert _sel([home, draw], played) == []      # un clic sur le 1 fait taire le marché


def test_other_markets_of_the_same_match_survive():
    """Jouer le 1X2 ne doit pas faire disparaître les totaux du même match."""
    h2h = _row(market="h2h", outcome="home")
    tot = _row(market="totals", outcome="over", line=2.5)
    played = {f"{h2h['event_key']}|h2h|None"}
    kept = _sel([h2h, tot], played)
    assert [b["market"] for b in kept] == ["totals"]


def test_bets_too_close_to_kickoff_are_excluded():
    """Même règle que les alertes : sous 15 min, ce sont surtout des lignes
    périmées."""
    assert _sel([_row(hours=0.1)]) == []
    assert len(_sel([_row(hours=0.5)])) == 1


def test_started_matches_are_excluded():
    assert _sel([_row(hours=-2)]) == []


def test_only_premium_grade_bets_are_kept():
    """/scan applique les regles du canal premium, pas celles du principal.

    Les 5-8 % d'EV du canal principal n'ont rien a faire dans cette liste —
    c'est le reproche exact de l'utilisateur."""
    assert _sel([_row(ev=5.0, odd=2.40)]) == []      # canal principal
    assert _sel([_row(ev=7.9, odd=2.40)]) == []      # juste sous les 8 %
    assert len(_sel([_row(ev=8.0, odd=2.40)])) == 1  # borne incluse
    assert _sel([_row(ev=15.0, odd=5.0)]) == []      # cote 4-6 sous les 20 %
    assert len(_sel([_row(ev=20.0, odd=5.0)])) == 1  # voie cotes hautes
    assert _sel([_row(ev=60.0, odd=14.0)]) == []     # hors bandes premium


def test_odds_bands_edges_are_inclusive():
    assert len(_sel([_row(ev=8.0, odd=1.5)])) == 1
    assert len(_sel([_row(ev=20.0, odd=6.0)])) == 1
    assert _sel([_row(ev=50.0, odd=1.49)]) == []
    assert _sel([_row(ev=50.0, odd=6.01)]) == []


# ------------------------------------------------------- déduplication ------

def test_same_selection_on_several_books_keeps_the_best_price():
    """Une sélection détectée sur trois books est UNE opportunité (§9), et on
    n'en joue qu'une — donc une seule ligne, au meilleur prix."""
    rows = [
        _row(book="unibet_be", odd=2.30, ev=10.0),
        _row(book="starcasino_sport", odd=2.45, ev=13.0),
        _row(book="circus_be", odd=2.38, ev=11.5),
    ]
    kept = _sel(rows)
    assert len(kept) == 1
    assert kept[0]["book"] == "starcasino_sport" and kept[0]["odd"] == 2.45


def test_results_are_sorted_by_ev_descending():
    rows = [_row(ev=8.5, outcome="home"), _row(ev=22.0, outcome="draw"),
            _row(ev=14.0, outcome="away")]
    assert [b["ev"] for b in _sel(rows)] == [22.0, 14.0, 8.5]


def test_unparseable_event_key_is_skipped_not_crashed():
    bad = _row()
    bad["event_key"] = "garbage"
    assert _sel([bad, _row(outcome="draw")]) != []


# ------------------------------------------------------------- marqueurs ----

@pytest.mark.parametrize("ev,odd,expected", [
    (12.0, 2.40, True),     # bande premium 1.5-4
    (25.0, 5.00, True),     # bande premium haute 4-6
    (60.0, 14.0, False),    # hors bandes : c'est du critique, pas du premium
    (6.0, 2.40, False),     # canal principal
    (15.0, 5.00, False),    # cotes 4-6 sous les 20 %
])
def test_is_premium_mirrors_the_premium_channel_rules(ev, odd, expected):
    assert is_premium(_cfg(), ev, odd) is expected


# --------------------------------------------------------------- rendu ------

def test_format_reports_emptiness_explicitly():
    """Un scan vide doit dire qu'il est vide. Le silence est indiscernable
    d'une panne — et c'est exactement ce qui s'est produit."""
    out = _texts([])
    assert len(out) == 1
    assert "Aucune value à jouer" in out[0]


def test_format_carries_everything_a_bet_needs():
    """Tout ce qu'il faut pour miser sans rouvrir autre chose : EV, book, match,
    date ET heure du coup d'envoi, pari, cote, cote juste, mise."""
    msg = _texts(_sel([_row(ev=12.4, odd=2.35, hours=3)]))[0]
    assert "+12.40% EV" in msg
    assert "Unibet" in msg                      # le book, en nom lisible
    assert "@ 2.35" in msg
    assert "(fair 2.09)" in msg                 # 2.35 / 1.124
    assert "17:00" in msg                       # heure de Bruxelles
    assert "Mise conseillée" in msg and "€" in msg


def test_format_shows_the_date_without_the_countdown():
    """La date suffit : « — dans 93 h » répète ce que la date porte déjà et
    allonge chaque ligne d'une liste triée par EV."""
    msg = _texts(_sel([_row(hours=93)]))[0]
    assert "📅" in msg
    assert "dans" not in msg


def test_format_separates_each_bet_with_a_blank_line():
    """Sans respiration entre les blocs, la liste est illisible sur téléphone."""
    msg = _texts(_sel([_row(outcome="home"), _row(outcome="draw")]))[0]
    assert "\n\n🎯" in msg


def test_format_derives_fair_odds_from_the_current_price():
    """La cote juste se déduit de la cote et de l'EV affichées, donc elle reste
    cohérente avec elles au lieu de dater de la première détection."""
    msg = _texts(_sel([_row(ev=25.0, odd=2.50)]))[0]
    assert "(fair 2.00)" in msg                 # 2.50 / 1.25


def test_format_names_the_selection_instead_of_home_away_draw():
    """« Anderlecht » plutôt que « home » : sur une liste, un label positionnel
    oblige à remonter à la ligne du match pour savoir de qui on parle."""
    home = _texts(_sel([_row(outcome="home")]))[0]
    assert "<b>Anderlecht</b>" in home and ">home<" not in home
    draw = _texts(_sel([_row(outcome="draw")]))[0]
    assert "Match nul" in draw
    tot = _row(market="totals", outcome="over", line=2.5)
    assert "Plus de 2.5" in _texts(_sel([tot]))[0]


def test_team_names_come_from_the_registry_not_the_raw_key(monkeypatch):
    """Le bug signalé : sans le registre, teams.display() retombe sur
    .capitalize() et rend « Clubbrugge » — collé et sans majuscule interne."""
    from src import teams

    monkeypatch.setitem(teams._DISPLAY, "clubbrugge", "Club Brugge")
    monkeypatch.setitem(teams._DISPLAY, "anderlecht", "Anderlecht")
    msg = _texts(_sel([_row()]))[0]
    assert "Anderlecht vs Club Brugge" in msg
    assert "Clubbrugge" not in msg


def test_format_splits_long_scans_across_messages():
    """Telegram refuse au-delà de 4096 caractères — un scan chargé doit sortir
    en plusieurs messages plutôt que d'échouer en silence."""
    rows = [_row(outcome=f"o{i}", ev=10.0 + i * 0.01) for i in range(60)]
    out = _texts(_sel(rows))
    assert len(out) > 1
    assert all(len(m) <= 4096 for m in out)


def test_format_escapes_html_in_team_names():
    """Un nom d'équipe contenant < ou & casserait le parse_mode HTML."""
    row = _row(home="A & <b>B</b>")
    msg = _texts(_sel([row]))[0]
    assert "&amp;" in msg and "&lt;b&gt;" in msg


# ------------------------------------------------------- fraîcheur réelle ---
# Une opportunité n'a qu'UNE ligne en base, écrite à la première détection.
# detected_at ne bouge jamais : c'est last_seen_at que le daemon rafraîchit à
# chaque cycle où il revoit le pari. Ces tests verrouillent cette distinction —
# s'en tromper cache exactement les paris qu'on veut voir.

def test_reseen_bet_refreshes_last_seen_without_touching_detection(tmp_path):
    from src.models import Book, MarketType, Outcome, ValueBet
    from src.storage import Storage

    st = Storage(tmp_path / "t.db")
    t0 = datetime(2026, 8, 4, 9, 0, tzinfo=timezone.utc)

    def vb(at, odd, ev):
        return ValueBet(
            event_key="209906010000::a__vs__b", book=Book.UNIBET_BE,
            market=MarketType.H2H, outcome=Outcome(label="home"),
            odd_taken=odd, fair_prob=0.55, fair_odd=1.80, ev_pct=ev,
            kelly_stake_pct=1.5, detected_at=at,
        )

    first = st.insert_value_bet(vb(t0, 2.40, 12.0))
    again = st.insert_value_bet(vb(t0 + timedelta(hours=3), 2.20, 6.0))
    assert again == first, "une ré-détection ne doit pas créer une seconde ligne"

    import sqlite3
    con = sqlite3.connect(str(tmp_path / "t.db"))
    con.row_factory = sqlite3.Row
    r = con.execute("SELECT * FROM value_bets WHERE id=?", (first,)).fetchone()
    con.close()
    # La détection est intacte : tout le CLV compare la clôture à l'EV de départ.
    assert r["detected_at"] == t0.isoformat()
    assert r["odd_taken"] == 2.40 and r["ev_pct"] == 12.0
    # La fraîcheur et le prix courant ont suivi.
    assert r["last_seen_at"] == (t0 + timedelta(hours=3)).isoformat()
    assert r["last_odd"] == 2.20 and r["last_ev"] == 6.0


def test_scan_selects_on_last_seen_not_detected_at(tmp_path, monkeypatch):
    """Le test qui aurait attrapé le bug, et il passe par fetch_playable — pas
    par une requête recopiée dans le test, qui n'aurait rien prouvé.

    Un pari détecté il y a 3 h mais revu il y a 10 s est jouable ; un pari
    détecté il y a 5 min et mort depuis ne l'est pas. Filtrer sur detected_at
    intervertit exactement les deux."""
    import sqlite3
    import bot_listener
    from src.storage import Storage

    db = tmp_path / "t.db"
    Storage(db)                            # schéma + migrations
    monkeypatch.setattr(bot_listener, "DB_PATH", db)
    monkeypatch.setattr("src.alerter._PLAYS_DB", db)

    now = datetime(2026, 8, 4, 12, 0, tzinfo=timezone.utc)
    iso = lambda m: (now - timedelta(minutes=m)).isoformat()  # noqa: E731
    kickoff = (now + timedelta(hours=3)).strftime("%Y%m%d%H%M")
    con = sqlite3.connect(str(db))
    con.executemany(
        "INSERT INTO value_bets(event_key, book, market, outcome_label, line, "
        "odd_taken, fair_prob, fair_odd, ev_pct, kelly_pct, detected_at, "
        "last_seen_at, last_odd, last_ev) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        [
            # vieille détection, toujours vivante -> doit sortir, au prix ACTUEL
            (f"{kickoff}::vivant__vs__x", "unibet_be", "h2h", "home", None,
             2.40, 0.55, 1.8, 12.0, 1.5, iso(180), iso(0.2), 2.35, 11.0),
            # détection récente, morte depuis -> ne doit pas sortir
            (f"{kickoff}::mort__vs__y", "unibet_be", "h2h", "home", None,
             2.40, 0.55, 1.8, 12.0, 1.5, iso(5), iso(40), 2.40, 12.0),
        ],
    )
    con.commit()
    con.close()

    bets = bot_listener.fetch_playable(_cfg(), now=now)
    assert [b["home"] for b in bets] == ["Vivant"]
    # Et c'est le dernier prix vu qui est proposé, pas celui d'il y a 3 h :
    # miser à 2.40 quand le book affiche 2.35 revient à courir après une cote
    # qui n'existe plus.
    assert bets[0]["odd"] == 2.35 and bets[0]["ev"] == 11.0


def test_scan_query_still_sees_rows_written_before_the_migration(tmp_path):
    """Les lignes déjà en base n'ont pas de last_seen_at — le COALESCE doit les
    rattraper plutôt que de les faire disparaître au redémarrage."""
    import sqlite3
    from src.storage import Storage

    Storage(tmp_path / "t.db")
    con = sqlite3.connect(str(tmp_path / "t.db"))
    now = datetime(2026, 8, 4, 12, 0, tzinfo=timezone.utc)
    con.execute(
        "INSERT INTO value_bets(event_key, book, market, outcome_label, line, "
        "odd_taken, fair_prob, fair_odd, ev_pct, kelly_pct, detected_at) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
        ("old", "unibet_be", "h2h", "home", None, 2.4, 0.55, 1.8, 12.0, 1.5,
         (now - timedelta(minutes=2)).isoformat()),
    )
    con.commit()
    n = con.execute(
        "SELECT COUNT(*) FROM value_bets WHERE COALESCE(last_seen_at, detected_at) >= ?",
        ((now - timedelta(minutes=10)).isoformat(),),
    ).fetchone()[0]
    con.close()
    assert n == 1


# ------------------------------------------------------------- dispatch -----
# Dans un CANAL Telegram, un message posté arrive en "channel_post" et non en
# "message". Les canaux d'alerte du projet en sont : n'écouter que "message"
# faisait que /scan ne recevait jamais rien, sans la moindre trace au journal.

def _dispatch_capture(monkeypatch, upd):
    import bot_listener
    seen = {}
    monkeypatch.setattr(bot_listener, "handle_message", lambda m: seen.setdefault("msg", m))
    monkeypatch.setattr(bot_listener, "handle_callback", lambda c: seen.setdefault("cb", c))
    bot_listener.dispatch(upd)
    return seen


def test_dispatch_routes_channel_post_like_a_message(monkeypatch):
    post = {"chat": {"id": -100123}, "text": "/scan"}
    assert _dispatch_capture(monkeypatch, {"update_id": 1, "channel_post": post}) == {"msg": post}


def test_dispatch_routes_plain_message(monkeypatch):
    msg = {"chat": {"id": 42}, "text": "/scan"}
    assert _dispatch_capture(monkeypatch, {"update_id": 1, "message": msg}) == {"msg": msg}


def test_dispatch_routes_callback_query(monkeypatch):
    cb = {"id": "x", "data": "play:abc"}
    assert _dispatch_capture(monkeypatch, {"update_id": 1, "callback_query": cb}) == {"cb": cb}


def test_dispatch_logs_unhandled_update_types(monkeypatch, capsys):
    _dispatch_capture(monkeypatch, {"update_id": 1, "poll_answer": {}})
    assert "poll_answer" in capsys.readouterr().out


def test_allowed_updates_covers_every_type_dispatch_handles():
    """Telegram ne livre que ce qui est demandé : un type géré par dispatch mais
    absent d'allowed_updates n'arriverait jamais — l'erreur d'origine."""
    import bot_listener
    for kind in ("callback_query", "message", "channel_post", "edited_channel_post"):
        assert kind in bot_listener.ALLOWED_UPDATES


# ------------------------------------------------------ chargement du .env --
# bot_listener est lance par systemd sans EnvironmentFile : sans chargement
# explicite, os.environ ne contient aucune config Telegram et /scan se
# desactive tout seul sur une installation qui marche. load_token() lisait deja
# .env pour son propre compte, ce qui masquait le probleme — le service
# demarrait normalement et ne repondait a rien.

def test_load_env_file_populates_os_environ(tmp_path, monkeypatch):
    from src.config import load_env_file

    env = tmp_path / ".env"
    env.write_text(
        "# un commentaire\n"
        "\n"
        "TELEGRAM_BOT_TOKEN=abc123\n"
        'TELEGRAM_PREMIUM_CHAT_ID="-1001234"\n'
        "LIGNE_SANS_EGAL\n"
    )
    for k in ("TELEGRAM_BOT_TOKEN", "TELEGRAM_PREMIUM_CHAT_ID"):
        monkeypatch.delenv(k, raising=False)
    assert load_env_file(env) == 2
    import os
    assert os.environ["TELEGRAM_BOT_TOKEN"] == "abc123"
    assert os.environ["TELEGRAM_PREMIUM_CHAT_ID"] == "-1001234"   # guillemets retires


def test_load_env_file_does_not_override_the_environment(monkeypatch, tmp_path):
    """Un override passe au service doit rester prioritaire sur le fichier."""
    from src.config import load_env_file
    import os

    env = tmp_path / ".env"
    env.write_text("TELEGRAM_CHAT_ID=du_fichier\n")
    monkeypatch.setenv("TELEGRAM_CHAT_ID", "de_l_environnement")
    load_env_file(env)
    assert os.environ["TELEGRAM_CHAT_ID"] == "de_l_environnement"


def test_load_env_file_tolerates_a_missing_file(tmp_path):
    from src.config import load_env_file
    assert load_env_file(tmp_path / "pas_la") == 0


def test_scan_accepts_every_configured_channel(monkeypatch, tmp_path):
    """La garde de chat doit connaitre TOUS les canaux du .env, pas seulement
    le principal — /scan est tape depuis le premium."""
    from src.config import load_env_file
    import bot_listener
    from src.alerter import TelegramConfig

    env = tmp_path / ".env"
    env.write_text(
        "TELEGRAM_BOT_TOKEN=t\n"
        "TELEGRAM_CHAT_ID=-100main\n"
        "TELEGRAM_PREMIUM_CHAT_ID=-100prem\n"
        "TELEGRAM_CRITICAL_CHAT_ID=-100crit\n"
        "TELEGRAM_SUREBET_CHAT_ID=-100sure\n"
        "TELEGRAM_CLV_CHAT_ID=-100clv\n"
    )
    for k in ("TELEGRAM_BOT_TOKEN", "TELEGRAM_CHAT_ID", "TELEGRAM_PREMIUM_CHAT_ID",
              "TELEGRAM_CRITICAL_CHAT_ID", "TELEGRAM_SUREBET_CHAT_ID",
              "TELEGRAM_CLV_CHAT_ID", "TELEGRAM_LIVE_SUREBET_CHAT_ID"):
        monkeypatch.delenv(k, raising=False)
    load_env_file(env)
    cfg = TelegramConfig.from_env()
    assert cfg is not None, "config illisible alors que .env est complet"
    assert bot_listener._allowed_chats(cfg) == {
        "-100main", "-100prem", "-100crit", "-100sure", "-100clv",
    }


# ------------------------------------------------- bouton « Tout jouer » ----
# Un clic engage TOUS les paris du message. C'est le comportement demande, et
# c'est aussi ce qui le rend delicat : le bouton doit couvrir exactement les
# paris affiches par SON message, et le scan suivant ne doit plus rien montrer.

def _scan_env(tmp_path, monkeypatch):
    """Une base neuve, avec bot_listener et l'alerter branches dessus."""
    import bot_listener
    from src.storage import Storage

    db = tmp_path / "t.db"
    Storage(db)
    monkeypatch.setattr(bot_listener, "DB_PATH", db)
    monkeypatch.setattr(bot_listener, "XLSX_PATH", tmp_path / "paris.xlsx")
    monkeypatch.setattr(bot_listener, "TRACK_PATH", tmp_path / "track.csv")
    monkeypatch.setattr("src.alerter._PLAYS_DB", db)
    return bot_listener, db


def test_scan_button_covers_exactly_its_own_message(tmp_path, monkeypatch):
    """Quand la liste est decoupee, un bouton par message, chacun sur ses paris.

    Un bouton qui vaudrait pour le scan entier enregistrerait des paris que son
    message ne montre meme pas."""
    bot_listener, _ = _scan_env(tmp_path, monkeypatch)
    bets = _sel([_row(outcome=f"o{i}", ev=10.0 + i * 0.01) for i in range(60)])
    parts = format_scan(bets, now=NOW)
    assert len(parts) > 1, "le cas interessant est la liste decoupee"
    assert sum(len(b) for _, b in parts) == len(bets)
    tokens = [bot_listener.register_scan(b, 1000.0) for _, b in parts]
    assert len(set(tokens)) == len(tokens)          # un jeton distinct par message
    for token, (_, part_bets) in zip(tokens, parts):
        assert len(bot_listener._scan_play_tokens(token)) == len(part_bets)


def test_playing_a_scan_empties_the_next_one(tmp_path, monkeypatch):
    """Le coeur de la demande, bout en bout : je scanne, j'appuie sur Tout
    jouer, je rescanne — il ne doit plus rien rester."""
    bot_listener, db = _scan_env(tmp_path, monkeypatch)
    sent = []
    monkeypatch.setattr(bot_listener, "tg", lambda m, **kw: sent.append((m, kw)) or {})

    rows = [_row(outcome="home", ev=12.0), _row(outcome="away", ev=9.0, odd=3.1),
            _row(home="Genk", away="Gent", ev=22.0, odd=5.2)]
    bets = _sel(rows)
    assert len(bets) == 3

    token = bot_listener.register_scan(bets, 1000.0)
    bot_listener.handle_scan_play({"id": "cb1", "data": f"scanplay:{token}",
                                   "message": {"chat": {"id": 1}, "message_id": 2}})

    # Deux cles de marche seulement : le home et le away du meme match sont le
    # meme marche, et c'est bien le marche entier qui est fait taire.
    from src.alerter import _load_played_keys
    _, played = _load_played_keys()
    assert len(played) == 2
    assert select_playable(rows, played, _cfg(), NOW) == []


def test_playing_a_scan_silences_the_other_outcomes_of_each_market(tmp_path, monkeypatch):
    """Jouer le 1 d'un 1X2 fait taire le X et le 2, comme pour les alertes."""
    bot_listener, _ = _scan_env(tmp_path, monkeypatch)
    monkeypatch.setattr(bot_listener, "tg", lambda m, **kw: {})

    home = _row(outcome="home", ev=12.0)
    bot_listener.handle_scan_play({
        "id": "cb", "data": f"scanplay:{bot_listener.register_scan(_sel([home]), 1000.0)}",
        "message": {"chat": {"id": 1}, "message_id": 2}})

    from src.alerter import _load_played_keys
    _, played = _load_played_keys()
    # Le X du meme marche n'etait pas dans le scan, il doit disparaitre aussi.
    assert select_playable([_row(outcome="draw", ev=12.0)], played, _cfg(), NOW) == []


def test_scan_button_is_idempotent_on_a_second_click(tmp_path, monkeypatch):
    """Telegram peut rejouer un callback ; un double clic ne doit pas doubler
    les lignes du tableur."""
    bot_listener, _ = _scan_env(tmp_path, monkeypatch)
    monkeypatch.setattr(bot_listener, "tg", lambda m, **kw: {})

    token = bot_listener.register_scan(_sel([_row()]), 1000.0)
    cb = {"id": "cb", "data": f"scanplay:{token}",
          "message": {"chat": {"id": 1}, "message_id": 2}}
    bot_listener.handle_scan_play(cb)
    bot_listener.handle_scan_play(cb)

    from openpyxl import load_workbook
    ws = load_workbook(bot_listener.XLSX_PATH).active
    assert ws.max_row == 2, "une seule ligne pour un pari clique deux fois"


def test_unknown_scan_token_answers_instead_of_crashing(tmp_path, monkeypatch):
    bot_listener, _ = _scan_env(tmp_path, monkeypatch)
    sent = []
    monkeypatch.setattr(bot_listener, "tg", lambda m, **kw: sent.append(kw) or {})
    bot_listener.handle_scan_play({"id": "cb", "data": "scanplay:inconnu",
                                   "message": {"chat": {"id": 1}, "message_id": 2}})
    assert any("introuvable" in str(k.get("text", "")) for k in sent)


def test_scanplay_callback_is_routed_to_its_handler(monkeypatch):
    """Le bouton du scan et celui des alertes partagent le canal callback_query :
    l'aiguillage doit distinguer scanplay: de play:."""
    import bot_listener
    seen = {}
    monkeypatch.setattr(bot_listener, "handle_scan_play", lambda cb: seen.setdefault("scan", cb))
    bot_listener.handle_callback({"id": "x", "data": "scanplay:abc"})
    assert seen.get("scan") is not None


def test_tg_drops_none_parameters(monkeypatch):
    """reply_markup=None partait en JSON null, que l'API refuse — et comme
    seul le scan VIDE n'a pas de bouton, lui seul disparaissait."""
    import bot_listener
    seen = {}

    class _Resp:
        status_code = 200
        @staticmethod
        def json():
            return {"ok": True}

    monkeypatch.setattr(bot_listener.requests, "post",
                        lambda url, json, timeout: seen.update(json) or _Resp())
    bot_listener.tg("sendMessage", chat_id="c", text="x", reply_markup=None)
    assert "reply_markup" not in seen
    assert seen["text"] == "x"


def test_tg_logs_a_telegram_refusal(monkeypatch, capsys):
    """Un envoi refuse ne laissait aucune trace : ni erreur, ni message."""
    import bot_listener

    class _Resp:
        status_code = 200
        @staticmethod
        def json():
            return {"ok": False, "description": "Bad Request: cant parse entities"}

    monkeypatch.setattr(bot_listener.requests, "post",
                        lambda url, json, timeout: _Resp())
    bot_listener.tg("sendMessage", chat_id="c", text="x")
    assert "cant parse entities" in capsys.readouterr().out


def test_empty_scan_still_sends_a_message(tmp_path, monkeypatch):
    """Bout en bout : base vide -> /scan doit quand meme repondre."""
    import bot_listener
    from src.storage import Storage

    db = tmp_path / "t.db"
    Storage(db)
    monkeypatch.setattr(bot_listener, "DB_PATH", db)
    monkeypatch.setattr("src.alerter._PLAYS_DB", db)
    monkeypatch.setenv("TELEGRAM_BOT_TOKEN", "t")
    monkeypatch.setenv("TELEGRAM_CHAT_ID", "42")
    sent = []
    monkeypatch.setattr(bot_listener, "tg", lambda m, **kw: sent.append((m, kw)) or {})

    bot_listener.handle_message({"chat": {"id": 42}, "text": "/scan"})
    assert [m for m, _ in sent] == ["sendMessage"]
    assert "Aucune value à jouer" in sent[0][1]["text"]


# ------------------------------------------------------------- /book --------
# Choisir de quels books on est NOTIFIÉ. La contrainte posée est stricte : un
# book décoché doit continuer d'être scrapé, stocké, suivi et exporté. Seul
# l'envoi Telegram disparaît.

def test_a_book_is_alerting_until_it_is_explicitly_muted(tmp_path):
    """Table d'exceptions, pas d'inscriptions : ajouter un scraper ne demande
    rien, et l'oubli d'une ligne ne rend jamais un book muet par accident."""
    from src.storage import Storage
    st = Storage(tmp_path / "t.db")
    assert st.books_alert_off() == set()
    assert st.toggle_book_alert("napoleon_be") is False      # -> coupé
    assert st.books_alert_off() == {"napoleon_be"}
    assert st.toggle_book_alert("napoleon_be") is True       # -> réactivé
    assert st.books_alert_off() == set()


def test_a_muted_book_sends_nothing(tmp_path, monkeypatch):
    from src.alerter import TelegramAlerter, TelegramConfig
    from src.storage import Storage
    db = tmp_path / "t.db"
    Storage(db).toggle_book_alert("unibet_be")
    monkeypatch.setattr("src.alerter._PLAYS_DB", db)

    from src.models import Book, MarketType, Outcome, ValueBet
    from datetime import datetime as _dt, timezone as _tz
    calls = []

    class FakeClient:
        def post(self, url, json):
            calls.append(json["chat_id"])
            class _R: status_code = 200
            return _R()
        def close(self): pass

    bet = ValueBet(event_key="209906010000::a__vs__b", book=Book.UNIBET_BE,
                   market=MarketType.H2H, outcome=Outcome(label="home"),
                   odd_taken=2.4, fair_prob=0.5, fair_odd=2.0, ev_pct=12.0,
                   kelly_stake_pct=1.5, detected_at=_dt.now(_tz.utc))
    cfg = TelegramConfig(bot_token="t", chat_id="c", min_ev_pct=3.0,
                         main_max_ev_pct=100.0, min_send_interval_s=0.0)
    with TelegramAlerter(cfg, client=FakeClient()) as a:
        assert a.send_value_bet(bet) is False
    assert calls == []


def test_muting_a_book_never_touches_collection(tmp_path):
    """Le point non négociable : la détection reste en base, avec sa cote et
    son EV. Seule la notification disparaît."""
    from src.models import Book, MarketType, Outcome, ValueBet
    from src.storage import Storage
    from datetime import datetime as _dt, timezone as _tz
    st = Storage(tmp_path / "t.db")
    st.toggle_book_alert("unibet_be")
    vb = ValueBet(event_key="209906010000::a__vs__b", book=Book.UNIBET_BE,
                  market=MarketType.H2H, outcome=Outcome(label="home"),
                  odd_taken=2.4, fair_prob=0.5, fair_odd=2.0, ev_pct=20.0,
                  kelly_stake_pct=1.5, detected_at=_dt.now(_tz.utc))
    vid = st.insert_value_bet(vb)
    import sqlite3
    con = sqlite3.connect(str(tmp_path / "t.db"))
    row = con.execute("SELECT book, ev_pct FROM value_bets WHERE id=?", (vid,)).fetchone()
    con.close()
    assert row == ("unibet_be", 20.0), "la détection doit rester intacte"


def test_the_keyboard_shows_the_state_of_each_book():
    from bot_listener import book_keyboard
    kb = book_keyboard(["unibet_be", "napoleon_be"], {"napoleon_be"})
    flat = [b for row in kb["inline_keyboard"] for b in row]
    by_data = {b["callback_data"]: b["text"] for b in flat}
    assert by_data["bookalert:unibet_be"].startswith("✅")
    assert by_data["bookalert:napoleon_be"].startswith("☐")
    assert "bookalert:__all__" in by_data and "bookalert:__none__" in by_data


def test_the_book_list_follows_what_actually_produced_detections(tmp_path):
    """Liste dynamique : un book ajouté apparaît seul, un book retiré
    disparaît, personne ne tient un second inventaire."""
    from src.models import Book, MarketType, Outcome, ValueBet
    from src.storage import Storage
    from datetime import datetime as _dt, timezone as _tz, timedelta as _td
    st = Storage(tmp_path / "t.db")
    now = _dt.now(_tz.utc)
    for i, (bk, when) in enumerate([(Book.UNIBET_BE, now),
                                    (Book.CIRCUS_BE, now - _td(days=30))]):
        st.insert_value_bet(ValueBet(
            event_key=f"20990601000{i}::a__vs__b", book=bk, market=MarketType.H2H,
            outcome=Outcome(label="home"), odd_taken=2.4, fair_prob=0.5,
            fair_odd=2.0, ev_pct=10.0, kelly_stake_pct=1.5, detected_at=when))
    assert st.books_seen(days=7) == ["unibet_be"], "Circus est trop ancien"
