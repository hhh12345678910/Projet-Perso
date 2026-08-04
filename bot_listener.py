#!/usr/bin/env python3
"""Ecoute les clics sur le bouton "Jouer" des alertes Telegram et enregistre
le pari dans un fichier Excel (data/paris.xlsx) + marque placed=1 en base.

Architecture :
  - L'alerte est envoyee par le daemon avec un bouton inline dont le
    callback_data vaut "play:<value_bet_id>".
  - Ce service fait du long-polling getUpdates (un SEUL process doit le faire).
  - Au clic : bouton -> "Joue", placed=1, ligne ajoutee dans l'Excel.

Dependances :  pip install requests openpyxl
Token : lu depuis la variable d'env TELEGRAM_BOT_TOKEN (ou le .env du projet).

Lancement manuel :
    source .venv/bin/activate
    python bot_listener.py
"""
from __future__ import annotations

import os
import sqlite3
import sys
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from src.matcher import parse_event_key

ROOT = Path(__file__).resolve().parent
DB_PATH = ROOT / "data" / "valuebet.db"
XLSX_PATH = ROOT / "data" / "paris.xlsx"
TRACK_PATH = ROOT / "data" / "paris_track.csv"
OFFSET_PATH = ROOT / "data" / ".telegram_offset"
BANKROLL = 1000.0  # capital de depart pour la colonne "Capital"
TRACK_STAKE_EUR = 25.0  # mise fictive constante, pour comparer des signaux

# --- colonnes de l'Excel (ordre = ordre d'ecriture) ---
HEADERS = [
    "Date", "Sport", "Match", "Book", "Selection",
    "Cote", "Mise EUR", "Resultat", "P&L EUR", "Capital", "EV %", "id",
]
COL = {name: chr(ord("A") + i) for i, name in enumerate(HEADERS)}  # A..L


def load_token() -> str:
    tok = os.environ.get("TELEGRAM_BOT_TOKEN", "").strip()
    if tok:
        return tok
    env = ROOT / ".env"
    if env.exists():
        for line in env.read_text().splitlines():
            line = line.strip()
            if line.startswith("TELEGRAM_BOT_TOKEN="):
                return line.split("=", 1)[1].strip().strip('"').strip("'")
    raise SystemExit("TELEGRAM_BOT_TOKEN introuvable (env ou .env)")


API = ""  # rempli dans main()


def tg(method: str, **params):
    r = requests.post(f"{API}/{method}", json=params, timeout=65)
    return r.json()


# ---------------------------------------------------------------- Excel -----
def ensure_workbook():
    if XLSX_PATH.exists():
        return
    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Paris"
    ws.append(HEADERS)
    wb.save(XLSX_PATH)


def already_logged(vb_id: int) -> bool:
    """Evite les doublons si on reclique sur le meme pari."""
    if not XLSX_PATH.exists():
        return False
    wb = load_workbook(XLSX_PATH)
    ws = wb.active
    idx = HEADERS.index("id")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and str(row[idx]) == str(vb_id):
            return True
    return False


def append_bet(bet: dict) -> None:
    ensure_workbook()
    wb = load_workbook(XLSX_PATH)
    ws = wb.active
    r = ws.max_row + 1  # ligne en cours d'ecriture

    pnl = COL["P&L EUR"] + str(r)
    res = COL["Resultat"] + str(r)
    cote = COL["Cote"] + str(r)
    mise = COL["Mise EUR"] + str(r)
    pnl_col = COL["P&L EUR"]

    # P&L auto selon le resultat saisi a la main (Gagne / Perdu / Annule)
    pnl_formula = (
        f'=IF({res}="Gagne",{mise}*({cote}-1),'
        f'IF({res}="Perdu",-{mise},'
        f'IF({res}="Annule",0,"")))'
    )
    # Capital = bankroll + somme des P&L jusqu'a cette ligne
    capital_formula = f"={BANKROLL}+SUM({pnl_col}$2:{pnl_col}{r})"

    ws.append([
        bet["date"], bet["sport"], bet["match"], bet["book"], bet["selection"],
        bet["cote"], bet["mise"], "", pnl_formula, capital_formula,
        bet["ev"], bet["id"],
    ])

    # menu deroulant Gagne/Perdu/Annule sur la colonne Resultat
    dv = DataValidation(type="list", formula1='"Gagne,Perdu,Annule"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws[res])

    wb.save(XLSX_PATH)


# -------------------------------------------------------------- Database ----
def fetch_bet(token: str) -> dict | None:
    """Lit la ligne pending_plays ecrite par l'alerter au moment de l'alerte."""
    con = sqlite3.connect(str(DB_PATH))
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA busy_timeout=5000")
    row = con.execute(
        "SELECT * FROM pending_plays WHERE token = ?", (token,)
    ).fetchone()
    con.close()
    if not row:
        return None
    return {
        "id": token,
        "date": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M"),
        "sport": row["sport"] or "",
        "match": row["match"],
        "book": row["book"],
        "selection": row["selection"],
        "cote": round(row["cote"], 3),
        "mise": round(row["mise"] or 0.0, 2),
        "ev": round(row["ev"], 1),
        "dedup_key": (row["dedup_key"] if "dedup_key" in row.keys() else None),
    }


def _record_played(bet: dict) -> None:
    """Enregistre le pari joue : suppression des alertes sur la selection, et
    une ligne dans le fichier de suivi.

    Le dedup_key vaut exactement event_key|market|outcome|line, ce qui suffit a
    retrouver le value_bet correspondant — et donc a rattacher plus tard la
    ligne de cloture, le CLV et le resultat a ce clic."""
    dedup_key = bet.get("dedup_key")
    if not dedup_key:
        return
    now = datetime.now(timezone.utc)

    # La suppression d'alerte d'abord, en sqlite nu : c'est elle qui compte, et
    # elle ne doit dependre d'aucun des enrichissements qui suivent. Si le
    # suivi casse, on perd une ligne de tableur ; si celui-ci casse, la
    # selection continue d'alerter apres avoir ete jouee.
    con = sqlite3.connect(str(DB_PATH))
    try:
        con.execute("PRAGMA busy_timeout=5000")
        con.execute(
            "CREATE TABLE IF NOT EXISTS played_bets (dedup_key TEXT PRIMARY KEY, played_at TEXT)"
        )
        con.execute(
            "INSERT OR IGNORE INTO played_bets(dedup_key, played_at) VALUES (?, ?)",
            (dedup_key, now.isoformat()),
        )
        con.commit()
    finally:
        con.close()

    try:
        _track_played(bet, dedup_key, now)
    except Exception as e:      # le suivi est informatif, jamais bloquant
        print(f"suivi non enregistre ({type(e).__name__}: {e})")


def _track_played(bet: dict, dedup_key: str, now: datetime) -> None:
    """Enrichit le clic (value bet d'origine, cote, EV, mise) et ajoute une
    ligne au fichier de suivi."""
    from src.storage import Storage
    from src import track

    storage = Storage(DB_PATH)
    parts = dedup_key.split("|")
    vb = None
    if len(parts) == 4:
        line = None if parts[3] in ("None", "") else float(parts[3])
        vb = storage.latest_value_bet_for(parts[0], parts[1], parts[2], line)

    # La ligne existe forcement : la suppression d'alerte vient de l'inserer.
    # Ce qui distingue un re-clic, c'est qu'elle porte deja une mise.
    existing = storage.played_bet(dedup_key)
    already = existing is not None and existing["stake"] is not None
    storage.record_played_bet(
        dedup_key=dedup_key, played_at=now, stake=track.STAKE_EUR,
        value_bet=vb, sport=bet.get("sport") or "", book=bet.get("book") or "",
        odd_taken=bet.get("cote"), ev_pct=bet.get("ev"),
    )
    if already:
        return  # deja suivi : ne pas dupliquer la ligne du fichier

    row = [""] * len(track.HEADERS)
    def put(col: str, value) -> None:
        row[track.HEADERS.index(col)] = value

    put("Date", now.strftime("%Y-%m-%d %H:%M"))
    put("Sport", bet.get("sport") or "")
    put("Match", bet.get("match") or "")
    put("Book", bet.get("book") or "")
    put("Sélection", bet.get("selection") or "")
    put("Cote prise", f"{float(bet['cote']):.2f}" if bet.get("cote") else "")
    put("EV %", f"{float(bet['ev']):.2f}" if bet.get("ev") is not None else "")
    put("Mise fictive", f"{track.STAKE_EUR:.2f}")
    if vb is not None:
        put("Marché", vb["market"])
        put("Cote fair (détection)", f"{float(vb['fair_odd']):.2f}")
        put("event_key", vb["event_key"])
    elif len(parts) == 4:
        put("Marché", parts[1])
        put("event_key", parts[0])
    # CLV, resultat et P&L restent vides : ils n'existent qu'apres le coup
    # d'envoi. `track-update` reconstruit le fichier complet a ce moment-la.
    track.append(TRACK_PATH, row)


# ------------------------------------------------------------- /scan --------
# Une détection plus vieille que ça n'est plus jouable : le daemon boucle en
# 15-20 s, donc une ligne de dix minutes a déjà été réévaluée des dizaines de
# fois. Si elle n'est pas réapparue, c'est que le prix a bougé.
SCAN_WINDOW_MIN = int(os.environ.get("SCAN_WINDOW_MIN", "10"))
SCAN_MAX_CHARS = 3500        # Telegram coupe à 4096 ; on garde de la marge
SCAN_MAX_BETS = 60           # au-delà, la liste n'est plus lisible de toute façon


def _team(normalised: str) -> str:
    """Nom lisible depuis un fragment de clé, via le même registre que les
    alertes. Retombe sur le fragment brut si le registre ne connaît pas encore
    l'équipe."""
    try:
        from src import teams
        return teams.display(normalised)
    except Exception:
        return normalised


def is_premium(cfg, ev: float, odd: float) -> bool:
    """Les règles du canal premium, et elles seules.

    Deux voies, sans plafond d'EV : cotes 1.5-4 à partir de 8 % d'EV, cotes 4-6
    à partir de 20 %. /scan est une liste de ce qui est jouable au sens du
    premium — pas du canal principal, dont les 5-8 % d'EV n'ont rien à y faire.
    """
    return bool(
        (ev >= cfg.min_premium_ev_pct
         and cfg.premium_min_odd <= odd <= cfg.premium_max_odd)
        or (ev >= cfg.premium_hi_min_ev
            and cfg.premium_hi_min_odd <= odd <= cfg.premium_hi_max_odd)
    )


def _selection_label(market: str, outcome: str, line, home: str, away: str) -> str:
    """« Anderlecht » plutôt que « home ». Sur une liste, un label positionnel
    oblige à remonter à la ligne du match pour savoir de qui on parle."""
    if market == "h2h":
        return {"home": home, "away": away, "draw": "Match nul"}.get(outcome, outcome)
    if line is not None:
        side = {"over": "Plus de", "under": "Moins de"}.get(outcome)
        if side:
            return f"{side} {line:g}"
        return f"{outcome} {line:g}"
    return outcome


def select_playable(rows, played_markets: set, cfg, now: datetime) -> list[dict]:
    """Parmi les détections récentes, celles encore jouables maintenant.

    Trois filtres, dans cet ordre :
      - le coup d'envoi est encore assez loin (même règle que les alertes) ;
      - le marché n'a pas déjà été joué — un clic sur « Jouer » sur le 1 d'un
        1X2 fait taire le X et le 2, comme pour les alertes ;
      - le pari atteindrait bien un canal.

    Puis une seule ligne par opportunité, au meilleur prix : la même sélection
    est détectée sur plusieurs books, mais on n'en joue qu'une (§9).
    """
    best: dict[tuple, dict] = {}
    for r in rows:
        parsed = parse_event_key(r["event_key"])
        if parsed is None:
            continue
        start = parsed[0]
        if (start - now).total_seconds() / 60 < cfg.min_minutes_to_kickoff:
            continue
        line = r["line"]
        if f"{r['event_key']}|{r['market']}|{line}" in played_markets:
            continue
        ev, odd = r["ev_pct"], r["odd_taken"]
        if not is_premium(cfg, ev, odd):
            continue
        key = (r["event_key"], r["market"], r["outcome_label"], line)
        prev = best.get(key)
        if prev is None or odd > prev["odd"]:
            # Noms via le registre d'équipes, exactement comme les alertes : la
            # clé ne porte que des fragments normalisés et collés
            # ("clubbrugge"), et teams.display() les rend lisibles.
            home, away = _team(parsed[1]), _team(parsed[2])
            # La cote juste se déduit de la cote et de l'EV courantes —
            # ev = (cote × p − 1) × 100, donc fair = cote / (1 + ev/100). Pas
            # besoin de la stocker, et elle reste cohérente avec le prix affiché
            # au lieu de dater de la première détection.
            fair = odd / (1.0 + ev / 100.0) if ev > -100 else None
            best[key] = {
                "ev": ev,
                "odd": odd,
                "fair": fair,
                "kelly": r["kelly_pct"],
                "book": r["book"],
                "market": r["market"],
                "outcome": r["outcome_label"],
                "line": line,
                "sport": (r["sport"] if "sport" in r.keys() else None) or "",
                "home": home,
                "away": away,
                "selection": _selection_label(
                    r["market"], r["outcome_label"], line, home, away
                ),
                "start": start,
            }
    out = sorted(best.values(), key=lambda b: -b["ev"])
    return out[:SCAN_MAX_BETS]


def fetch_playable(cfg, *, now: datetime | None = None) -> list[dict]:
    now = now or datetime.now(timezone.utc)
    since = (now - timedelta(minutes=SCAN_WINDOW_MIN)).isoformat()
    con = sqlite3.connect(str(DB_PATH))
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA busy_timeout=5000")
    try:
        # last_seen_at, pas detected_at : une opportunité n'a qu'une ligne,
        # écrite à la première détection et jamais redatée. Filtrer sur
        # detected_at ne montrerait que les paris NOUVELLEMENT apparus, en
        # cachant tous ceux qui tiennent depuis des heures — l'inverse de ce
        # qu'on veut. COALESCE couvre les lignes écrites avant la migration.
        # Le prix affiché est le dernier vu, pas celui de la détection : c'est
        # à celui-là qu'on peut miser maintenant.
        rows = con.execute(
            "SELECT v.event_key, v.book, v.market, v.outcome_label, v.line, "
            "       v.kelly_pct, "
            "       COALESCE(v.last_odd, v.odd_taken) AS odd_taken, "
            "       COALESCE(v.last_ev, v.ev_pct)     AS ev_pct, "
            "       e.sport "
            "FROM value_bets v LEFT JOIN events e ON e.event_key = v.event_key "
            "WHERE COALESCE(v.last_seen_at, v.detected_at) >= ?",
            (since,),
        ).fetchall()
    finally:
        con.close()
    from src.alerter import _load_played_keys
    _, played_markets = _load_played_keys()
    return select_playable(rows, played_markets, cfg, now)


def _esc(s: str) -> str:
    return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def format_scan(bets: list[dict], *, now: datetime | None = None,
                bankroll: float = 1000.0) -> list[str]:
    """Rend le scan en un ou plusieurs messages HTML sous la limite Telegram.

    Même disposition qu'une alerte — EV et book, match, coup d'envoi, pari avec
    la cote juste, mise conseillée — mais empilée en liste, une ligne vide entre
    chaque pari pour que l'œil sépare les blocs.
    """
    now = now or datetime.now(timezone.utc)
    if not bets:
        return [
            "🔎 <b>SCAN</b> — aucune value jouable actuellement.\n\n"
            f"<i>Critères premium : EV ≥ 8 % sur cotes 1.5-4, EV ≥ 20 % sur "
            f"cotes 4-6. Sont exclus les paris déjà joués, ceux à moins de "
            f"15 min du coup d'envoi, et ceux qui n'ont plus été vus depuis "
            f"{SCAN_WINDOW_MIN} min.</i>"
        ]
    plural = "s" if len(bets) > 1 else ""
    header = f"🔎 <b>SCAN</b> — {len(bets)} value{plural} jouable{plural}\n"
    blocks = []
    for b in bets:
        emoji = _SCAN_SPORT_EMOJI.get(b["sport"], "")
        fair = f" (fair {b['fair']:.2f})" if b.get("fair") else ""
        stake = _stake_line(b["ev"], b.get("kelly"), bankroll)
        blocks.append(
            f"\n🎯 <b>+{b['ev']:.2f}% EV</b> — {_esc(_book_label(b['book']))}\n"
            f"{emoji} {_esc(b['home'])} vs {_esc(b['away'])}\n"
            f"📅 {_esc(_kickoff(b['start'], now))}\n"
            f"Pari : <b>{_esc(b['selection'])}</b> @ {b['odd']:.2f}{fair}\n"
            f"{stake}\n"
        )
    msgs, cur = [], header
    for block in blocks:
        if len(cur) + len(block) > SCAN_MAX_CHARS:
            msgs.append(cur)
            cur = ""
        cur += block
    if cur.strip():
        msgs.append(cur)
    return msgs


_SCAN_SPORT_EMOJI = {"soccer": "⚽", "football": "⚽", "tennis": "🎾",
                     "basketball": "🏀", "hockey": "🏒"}


def _kickoff(start: datetime, now: datetime) -> str:
    """« Aujourd'hui 20:45 — dans 3 h », en heure de Bruxelles comme les
    alertes : une liste sans horaire oblige à rouvrir chaque match pour savoir
    lequel part en premier."""
    try:
        from src.alerter import _format_kickoff, _time_to_kickoff
        return f"{_format_kickoff(start, now)}{_time_to_kickoff(start, now)}"
    except Exception:
        return start.strftime("%d/%m %H:%M")


def _stake_line(ev: float, kelly: float | None, bankroll: float) -> str:
    """La mise, calculée comme dans les alertes — même mode (kelly ou flat),
    même arrondi. Une mise différente entre l'alerte et le scan pour le même
    pari serait pire que pas de mise du tout."""
    try:
        from src.alerter import _advised_stake_line
        return _advised_stake_line(ev, kelly, bankroll) or "Mise conseillée : —"
    except Exception:
        return "Mise conseillée : —"


def _book_label(book_value: str) -> str:
    try:
        from src.alerter import _BOOK_NAMES
        from src.models import Book
        return _BOOK_NAMES.get(Book(book_value), book_value)
    except Exception:
        return book_value


def _allowed_chats(cfg) -> set[str]:
    """Les chats du projet. Le bot est joignable par n'importe qui sur Telegram :
    sans cette garde, un inconnu qui trouve son nom obtiendrait la liste des
    paris en tapant /scan."""
    ids = (
        cfg.chat_id, cfg.surebet_chat_id, cfg.live_surebet_chat_id,
        cfg.clv_chat_id, cfg.critical_chat_id, cfg.premium_chat_id,
    )
    return {str(i) for i in ids if i}


def handle_message(msg: dict) -> None:
    text = (msg.get("text") or msg.get("caption") or "").strip()
    chat_id = str((msg.get("chat") or {}).get("id") or "")
    if not text.startswith("/"):
        return
    # "/scan@mon_bot arg" -> "/scan"
    cmd = text.split()[0].split("@", 1)[0].lower()
    if cmd not in ("/scan", "/start"):
        print(f"[{datetime.now():%H:%M:%S}] commande inconnue {cmd!r} (chat {chat_id})")
        return

    from src.alerter import TelegramConfig
    cfg = TelegramConfig.from_env()
    if cfg is None:
        print("TELEGRAM_BOT_TOKEN/CHAT_ID absents — /scan desactive")
        return
    allowed = _allowed_chats(cfg)
    if chat_id not in allowed:
        # Dire lequel, et contre quoi : sans ça, « /scan ne répond pas » et
        # « le bot n'a rien reçu » sont indiscernables.
        print(f"[{datetime.now():%H:%M:%S}] {cmd} refuse — chat {chat_id} "
              f"absent des TELEGRAM_*_CHAT_ID connus ({sorted(allowed)})")
        return

    if cmd == "/start":
        tg("sendMessage", chat_id=chat_id, parse_mode="HTML",
           text="Commandes :\n/scan — les value bets encore jouables maintenant")
        return

    try:
        bets = fetch_playable(cfg)
    except Exception as e:
        tg("sendMessage", chat_id=chat_id, text=f"Scan indisponible ({type(e).__name__})")
        print(f"scan echoue: {type(e).__name__}: {e}")
        return
    for part in format_scan(bets, bankroll=cfg.bankroll):
        tg("sendMessage", chat_id=chat_id, parse_mode="HTML",
           text=part, disable_web_page_preview=True)
    print(f"[{datetime.now():%H:%M:%S}] /scan -> {len(bets)} paris (chat {chat_id})")


# ----------------------------------------------------------- Callbacks ------
def handle_callback(cb: dict) -> None:
    data = cb.get("data", "")
    if not data.startswith("play:"):
        return
    cb_id = cb["id"]
    token = data.split(":", 1)[1]
    if not token:
        tg("answerCallbackQuery", callback_query_id=cb_id, text="Jeton invalide")
        return

    if already_logged(token):
        tg("answerCallbackQuery", callback_query_id=cb_id, text="Deja enregistre")
        _mark_button_done(cb)
        return

    bet = fetch_bet(token)
    if not bet:
        tg("answerCallbackQuery", callback_query_id=cb_id, text="Pari introuvable")
        return

    append_bet(bet)
    _record_played(bet)
    _mark_button_done(cb)
    tg("answerCallbackQuery", callback_query_id=cb_id,
       text=f"Enregistre : {bet['mise']:.2f} EUR @ {bet['cote']}")
    print(f"[{datetime.now():%H:%M:%S}] logged {token} {bet['match']} {bet['book']}")


def _mark_button_done(cb: dict) -> None:
    """Remplace le bouton par un '✅ Joue' non cliquable (effet 'vert')."""
    msg = cb.get("message", {})
    chat_id = msg.get("chat", {}).get("id")
    msg_id = msg.get("message_id")
    if chat_id is None or msg_id is None:
        return
    tg("editMessageReplyMarkup", chat_id=chat_id, message_id=msg_id,
       reply_markup={"inline_keyboard": [[{"text": "✅ Joue", "callback_data": "noop"}]]})


# --------------------------------------------------------------- Loop -------
# channel_post en plus de message : dans un CANAL Telegram, un message posté
# n'arrive pas en "message" mais en "channel_post". Les canaux d'alerte de ce
# projet en sont, donc sans lui /scan ne recevait rien — et un update jamais
# reçu ne laisse aucune trace nulle part.
ALLOWED_UPDATES = ["callback_query", "message", "channel_post", "edited_channel_post"]


def dispatch(upd: dict) -> None:
    """Aiguille un update. Un type non traité est journalisé, jamais jeté en
    silence : sans ça « la commande ne répond pas » et « le bot n'a rien reçu »
    sont indiscernables (§13.12)."""
    if "callback_query" in upd:
        handle_callback(upd["callback_query"])
        return
    for kind in ("message", "channel_post", "edited_channel_post"):
        if kind in upd:
            handle_message(upd[kind])
            return
    print(f"update ignore: {[k for k in upd if k != 'update_id']}")


def read_offset() -> int:
    if OFFSET_PATH.exists():
        try:
            return int(OFFSET_PATH.read_text().strip())
        except ValueError:
            return 0
    return 0


def save_offset(off: int) -> None:
    OFFSET_PATH.write_text(str(off))


def main() -> None:
    global API
    API = f"https://api.telegram.org/bot{load_token()}"
    # Sous systemd, stdout n'est pas un terminal : Python le tamponne par blocs
    # de 4 Ko et les diagnostics n'arrivent au journal que bien plus tard, quand
    # on en a justement besoin tout de suite.
    try:
        sys.stdout.reconfigure(line_buffering=True)
    except Exception:
        pass
    # systemd lance ce service sans EnvironmentFile : sans ce chargement,
    # os.environ est vide de toute config Telegram et TelegramConfig.from_env()
    # renvoie None — /scan se desactivait tout seul sur une installation
    # parfaitement configuree. load_token() lisait deja .env pour son propre
    # compte, ce qui masquait le probleme : le bot demarrait normalement.
    from src.config import load_env_file
    print(f"env : {load_env_file(ROOT / '.env')} cles chargees depuis .env")

    # Sans ce branchement, teams.display() retombe sur .capitalize() et rend
    # « Clubbrugge » au lieu de « Club Brugge » : la cle d'evenement ne porte
    # que des fragments normalises et colles, c'est le registre qui sait les
    # reecrire. Le daemon l'initialise de son cote, pas ce service.
    try:
        from src import teams
        from src.storage import Storage
        teams.init(Storage(DB_PATH))
        print(f"registre d'equipes : {len(teams._DISPLAY)} noms charges")
    except Exception as e:
        print(f"registre d'equipes indisponible ({type(e).__name__}) — noms bruts")

    offset = read_offset()
    print(f"bot_listener demarre (offset={offset}, db={DB_PATH}, xlsx={XLSX_PATH})")
    try:
        from src.alerter import TelegramConfig
        _cfg = TelegramConfig.from_env()
        print(f"chats acceptes pour /scan : {sorted(_allowed_chats(_cfg)) if _cfg else 'AUCUN'}")
    except Exception as e:
        print("config Telegram illisible:", e)
    try:    # fait apparaitre /scan dans le menu Telegram ; sans effet sur le reste
        tg("setMyCommands", commands=[
            {"command": "scan", "description": "Les value bets encore jouables"},
        ])
    except Exception as e:
        print("setMyCommands ignore:", e)
    while True:
        try:
            resp = tg("getUpdates", offset=offset, timeout=50,
                      allowed_updates=ALLOWED_UPDATES)
            for upd in resp.get("result", []):
                offset = upd["update_id"] + 1
                dispatch(upd)
                save_offset(offset)
        except requests.RequestException as e:
            print("reseau:", e)
            time.sleep(3)
        except Exception as e:  # ne jamais laisser le service tomber
            print("erreur:", e)
            time.sleep(3)


if __name__ == "__main__":
    main()
