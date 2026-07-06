from __future__ import annotations

import os
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import replace
from datetime import datetime, timezone
from typing import Callable, Iterable

import time

import httpx
import typer
from rich.console import Console
from rich.table import Table

from .config import ScanConfig
from .devig import devig
from .ev import ev_pct, fair_odd, kelly_fraction, kelly_stake
from .matcher import parse_event_key, reconcile_event_keys
from .models import Book, FairLine, MarketType, OddQuote, Outcome, ValueBet
from .scrapers.betano import BetanoAuthError, BetanoScraper, parse_overview as betano_parse_overview
from .scrapers.betcenter import BetCenterScraper
from .scrapers.betfirst import BetFirstScraper, parse_events_table as betfirst_parse_events_table
from .scrapers.goldenpalace import GoldenPalaceScraper, parse_get_events as goldenpalace_parse_get_events
from .scrapers.ladbrokes import LadbrokesScraper, parse_prematch as ladbrokes_parse_prematch
from .scrapers.pinnacle import PinnacleScraper
from .scrapers.sevenelevenbe import SevenElevenScraper, parse_listview as sevenelevenbe_parse_listview
from .scrapers.bingoal import BingoalScraper, parse_listview as bingoal_parse_listview
from .scrapers.scooore import ScoooreScraper, parse_listview as scooore_parse_listview
from .scrapers.meridianbet import MeridianScraper, parse_offer as meridian_parse_offer
from .scrapers.napoleon import NapoleonScraper, parse_by_date as napoleon_parse_by_date
from .scrapers.starcasinosport import StarCasinoSportScraper, parse_get_events as starcasinosport_parse_get_events
from .scrapers.unibet import UnibetScraper, parse_listview as unibet_parse_listview
from .storage import Storage
from .surebet import find_surebets, Surebet
from .middle import find_middles, Middle
from .clv import (
    aggregate as clv_aggregate,
    clv_pct,
    event_started,
    group_by as clv_group_by,
    index_quotes_by_market,
)
from .alerter import TelegramConfig, send_alerts, send_surebet_alerts, send_clv_alerts, send_middle_alerts
from . import teams


app = typer.Typer(add_completion=False)
console = Console()


def _group_quotes(quotes: Iterable[OddQuote]) -> dict[tuple[str, MarketType, float | None], list[OddQuote]]:
    """Group Pinnacle quotes into competing-outcome sets per (event, market, line)."""
    groups: dict[tuple[str, MarketType, float | None], list[OddQuote]] = defaultdict(list)
    for q in quotes:
        groups[(q.event_key, q.market, q.outcome.line)].append(q)
    return groups


def _devig_group(group: list[OddQuote], method: str) -> dict[str, float] | None:
    """Run a devig on one (event, market, line) group's odds. Returns the
    label -> fair probability map, or None if the group is too thin or
    numerically degenerate."""
    if len(group) < 2:
        return None
    try:
        probs = devig([q.decimal_odd for q in group], method=method)
    except Exception:
        return None
    return {q.outcome.label: p for q, p in zip(group, probs)}


def build_fair_lines(
    pinnacle_quotes: list[OddQuote],
    method: str,
    *,
    secondary_quotes: list[OddQuote] | None = None,
    primary_weight: float = 0.7,
) -> dict[tuple[str, MarketType, float | None], FairLine]:
    """Build fair lines from Pinnacle, optionally blending with a secondary
    sharp source (Smarkets exchange) to cross-validate. Pinnacle keeps the
    higher weight by default because its volume and stability are higher;
    Smarkets fills in events Pinnacle doesn't price and gently pulls the
    estimate where the two disagree."""
    primary_groups = _group_quotes(pinnacle_quotes)
    secondary_groups = _group_quotes(secondary_quotes or [])

    fair: dict[tuple[str, MarketType, float | None], FairLine] = {}
    now = datetime.now(timezone.utc)
    for key in primary_groups.keys() | secondary_groups.keys():
        event_key_, market, line = key
        pin_probs = _devig_group(primary_groups.get(key, []), method)
        sec_probs = _devig_group(secondary_groups.get(key, []), method)

        if pin_probs and sec_probs:
            # Weighted blend on outcomes that both sources price; fall back to
            # the available source when only one carries a given label.
            blended: dict[str, float] = {}
            for label in set(pin_probs) | set(sec_probs):
                p, s = pin_probs.get(label), sec_probs.get(label)
                if p is not None and s is not None:
                    blended[label] = primary_weight * p + (1 - primary_weight) * s
                else:
                    blended[label] = p if p is not None else s
            # Renormalise so the row still sums to 1 after the blend.
            total = sum(blended.values())
            if total > 0:
                blended = {k: v / total for k, v in blended.items()}
            outcomes = blended
            ref_book = Book.PINNACLE
        elif pin_probs:
            outcomes = pin_probs
            ref_book = Book.PINNACLE
        else:
            outcomes = sec_probs  # type: ignore[assignment]
            ref_book = Book.SMARKETS

        if not outcomes:
            continue
        fair[key] = FairLine(
            event_key=event_key_,
            market=market,
            outcomes=outcomes,
            method=method,
            reference_book=ref_book,
            computed_at=now,
        )
    return fair


def find_value_bets(
    candidate_quotes: list[OddQuote],
    fair_lines: dict[tuple[str, MarketType, float | None], FairLine],
    cfg: ScanConfig,
) -> list[ValueBet]:
    # Pre-count distinct outcome labels per (event, market, line, book).
    # If a soft book offers fewer outcomes than Pinnacle's fair line (e.g.
    # hockey 2-way OT-included on Pinnacle vs 3-way regulation 1X2 on soft
    # books), the markets are structurally incompatible and must be skipped.
    from collections import defaultdict
    _soft_labels: dict[tuple, set[str]] = defaultdict(set)
    for q in candidate_quotes:
        if q.book != Book.PINNACLE:
            _soft_labels[(q.event_key, q.market, q.outcome.line, q.book)].add(q.outcome.label)
    soft_outcome_counts = {k: len(v) for k, v in _soft_labels.items()}

    out: list[ValueBet] = []
    now = datetime.now(timezone.utc)
    for q in candidate_quotes:
        if q.book == Book.PINNACLE:
            continue
        # Handicap markets are excluded for now: line semantics vary across
        # books (Pinnacle signs each side, e.g. home -1.0 / away +1.0, while
        # soft books carry both sides at the same |line|). That mismatch pairs
        # non-complementary lines in the devig and surfaces phantom value bets
        # (huge "EV" on away -1.0 etc.). Surebets already skip handicaps for the
        # same reason; mirror that here until per-book line conventions are
        # normalised.
        if q.market == MarketType.HANDICAP:
            continue
        fl = fair_lines.get((q.event_key, q.market, q.outcome.line))
        if fl is None:
            continue
        # Skip if the soft book doesn't offer the same number of outcomes as
        # the Pinnacle fair line (market structure mismatch).
        n_soft = soft_outcome_counts.get((q.event_key, q.market, q.outcome.line, q.book), 0)
        if n_soft != len(fl.outcomes):
            continue
        p = fl.outcomes.get(q.outcome.label)
        if p is None or p <= 0 or p >= 1:
            continue
        ev = ev_pct(q.decimal_odd, p)
        if ev < cfg.min_ev_pct or ev > cfg.max_ev_pct:
            continue
        out.append(ValueBet(
            event_key=q.event_key,
            book=q.book,
            market=q.market,
            outcome=q.outcome,
            odd_taken=q.decimal_odd,
            fair_prob=p,
            fair_odd=fair_odd(p),
            ev_pct=ev,
            kelly_stake_pct=kelly_fraction(q.decimal_odd, p) * cfg.kelly_fraction * 100.0,
            detected_at=now,
        ))
    return out


# Books that share a single odds feed (Kambi): Unibet and 711 price identically,
# so the same value bet on both is one opportunity, not two. UNIBET is the
# canonical book kept for storage/dedup; 711 rides along in `also_books`.
_TWIN_BOOK_GROUPS: tuple[tuple[Book, ...], ...] = (
    (Book.UNIBET_BE, Book.SEVEN_ELEVEN_BE, Book.BINGOAL_BE, Book.SCOOORE_BE),
)
_TWIN_PRIMARY = {grp: grp[0] for grp in _TWIN_BOOK_GROUPS}
_TWIN_OF = {b: grp for grp in _TWIN_BOOK_GROUPS for b in grp}


def merge_twin_book_value_bets(bets: list[ValueBet]) -> list[ValueBet]:
    """Collapse identical value bets coming from twin books (same Kambi feed,
    same price) into a single alert that names every book. Non-twin bets and
    twin bets that don't have a same-priced sibling pass through untouched."""
    twins: dict[tuple, list[ValueBet]] = defaultdict(list)
    out: list[ValueBet] = []
    for b in bets:
        if b.book in _TWIN_OF:
            key = (b.event_key, b.market, b.outcome.label, b.outcome.line,
                   round(b.odd_taken, 4), _TWIN_OF[b.book])
            twins[key].append(b)
        else:
            out.append(b)

    for key, group in twins.items():
        twin_group = key[5]
        primary_book = _TWIN_PRIMARY[twin_group]
        books_present = {b.book for b in group}
        # Keep the primary book's record if present, else the first seen.
        base = next((b for b in group if b.book == primary_book), group[0])
        extras = tuple(b for b in twin_group if b in books_present and b != base.book)
        out.append(replace(base, also_books=extras))
    return out


_OPPOSITE_OUTCOME = {"home": "away", "away": "home"}


def _flip_outcome_for_swap(outcome: Outcome, market: MarketType) -> Outcome:
    """When the matcher had to swap home/away to align a soft-book event_key
    with the Pinnacle reference, any outcome labels carried by quotes from
    that event are now pointing at the wrong team in the reference frame.
    Flip home↔away (draw stays); the totals over/under labels are
    team-symmetric so they pass through unchanged."""
    if market == MarketType.TOTALS:
        return outcome
    flipped_label = _OPPOSITE_OUTCOME.get(outcome.label, outcome.label)
    return replace(outcome, label=flipped_label)


def remap_to_reference(
    soft_quotes: list[OddQuote],
    reference_keys: Iterable[str],
) -> list[OddQuote]:
    """Re-key soft-book quotes onto the matching Pinnacle event_key via fuzzy
    matching, so they line up with the fair lines. When the matcher detects
    that the candidate listed the teams in swapped order (e.g. soft book has
    'Senegal vs Nigeria' while Pinnacle has 'Nigeria vs Senegal'), the home
    /away outcome labels are flipped on the way out so the rest of the
    pipeline compares apples to apples. Unmatched quotes are dropped."""
    soft_to_ref = reconcile_event_keys(
        reference_keys=list(reference_keys),
        candidate_keys={q.event_key for q in soft_quotes},
    )
    out: list[OddQuote] = []
    for q in soft_quotes:
        match = soft_to_ref.get(q.event_key)
        if match is None:
            continue
        ref_key, swap = match
        flipped_outcome = _flip_outcome_for_swap(q.outcome, q.market) if swap else q.outcome
        if ref_key == q.event_key and not swap:
            out.append(q)
        else:
            out.append(replace(q, event_key=ref_key, outcome=flipped_outcome))
    return out


def canonicalize_for_surebets(
    pinnacle_q: list[OddQuote],
    soft_raw: list[OddQuote],
) -> list[OddQuote]:
    """Re-key every quote (Pinnacle + soft books) onto a unified canonical key
    set so surebets can be found across books even on events Pinnacle does NOT
    price.

    Unlike remap_to_reference (which anchors on Pinnacle and drops anything
    Pinnacle doesn't list), this lets soft books anchor each other: Pinnacle
    keys seed the canonical set when present (cleanest team names), then each
    soft book is reconciled one at a time against the growing reference. The
    first book to price an event Pinnacle lacks becomes that event's anchor,
    and later books fuzzy-match onto it. Quotes that match adopt the anchor key
    (home/away flipped when the match was swapped); unmatched events seed new
    anchors so the next book can still align with them.

    This is for surebet detection only — value bets still need a Pinnacle fair
    line, so they keep using remap_to_reference."""
    canonical: list[OddQuote] = list(pinnacle_q)  # Pinnacle keeps its own keys
    ref_keys: set[str] = {q.event_key for q in pinnacle_q}

    # Reconcile each book as a unit so a book never matches against itself.
    by_book: dict[Book, list[OddQuote]] = defaultdict(list)
    for q in soft_raw:
        by_book[q.book].append(q)

    for _book, quotes in by_book.items():
        mapping = reconcile_event_keys(
            reference_keys=list(ref_keys),
            candidate_keys={q.event_key for q in quotes},
        )
        new_anchor_keys: set[str] = set()
        for q in quotes:
            match = mapping.get(q.event_key)
            if match is None:
                # No match anywhere yet — this event becomes its own anchor so
                # subsequent books can align onto it.
                canonical.append(q)
                new_anchor_keys.add(q.event_key)
                continue
            ref_key, swap = match
            flipped = _flip_outcome_for_swap(q.outcome, q.market) if swap else q.outcome
            if ref_key == q.event_key and not swap:
                canonical.append(q)
            else:
                canonical.append(replace(q, event_key=ref_key, outcome=flipped))
        ref_keys |= new_anchor_keys
    return canonical


def fetch_betano_quotes(betano_file: str | None = None) -> list[OddQuote]:
    """Parse Betano data. If `betano_file` is given, load the JSON from disk
    (the response body the user captured in their browser) instead of doing a
    live fetch — Cloudflare+DataDome bind cookies to the browser IP, so the
    live path only works from that machine. Returns [] with a warning if no
    file is given and the cookie is missing/expired."""
    if betano_file:
        import json as _json
        from pathlib import Path as _Path
        try:
            data = _json.loads(_Path(betano_file).read_text())
        except (OSError, ValueError) as e:
            console.print(f"[yellow]Betano file unreadable:[/yellow] {e}")
            return []
        return list(betano_parse_overview(data))
    try:
        with BetanoScraper() as bet:
            data = bet.fetch_live_overview()
        return list(betano_parse_overview(data))
    except BetanoAuthError as e:
        console.print(f"[yellow]Betano skipped:[/yellow] {e}")
        return []


def fetch_unibet_quotes(sport: str) -> list[OddQuote]:
    """Fetch + parse every Unibet (Kambi) event by iterating each leaf termKey
    under the sport's group — far better coverage than the bare listView."""
    try:
        with UnibetScraper() as uni:
            data = uni.fetch_all_events(sport)
        return list(unibet_parse_listview(data))
    except httpx.HTTPError as e:
        console.print(f"[yellow]Unibet skipped:[/yellow] {e}")
        return []


def fetch_sevenelevenbe_quotes(sport: str) -> list[OddQuote]:
    """Fetch + parse every 711 (Kambi) event — same coverage strategy as Unibet,
    just a different operator code on the shared Kambi offering API."""
    try:
        with SevenElevenScraper() as se:
            data = se.fetch_all_events(sport)
        return list(sevenelevenbe_parse_listview(data))
    except httpx.HTTPError as e:
        console.print(f"[yellow]711 skipped:[/yellow] {e}")
        return []


def fetch_bingoal_quotes(sport: str) -> list[OddQuote]:
    """Fetch + parse every Bingoal (Kambi) event — same coverage strategy as
    Unibet/711, just a different operator code on the shared Kambi offering API."""
    try:
        with BingoalScraper() as bg:
            data = bg.fetch_all_events(sport)
        return list(bingoal_parse_listview(data))
    except httpx.HTTPError as e:
        console.print(f"[yellow]Bingoal skipped:[/yellow] {e}")
        return []


def fetch_meridian_quotes(sport: str) -> list[OddQuote]:
    """Fetch + parse the MeridianBet prematch offer for a sport (own REST API,
    independent odds — genuinely widens value/surebet coverage)."""
    try:
        with MeridianScraper() as mb:
            data = mb.fetch_all_events(sport)
        return list(meridian_parse_offer(data))
    except httpx.HTTPError as e:
        console.print(f"[yellow]MeridianBet skipped:[/yellow] {e}")
        return []


def fetch_scooore_quotes(sport: str) -> list[OddQuote]:
    """Fetch + parse every Scooore (Kambi) event — same coverage strategy as
    Unibet/711/Bingoal, just a different operator code (bnlbe)."""
    try:
        with ScoooreScraper() as sc:
            data = sc.fetch_all_events(sport)
        return list(scooore_parse_listview(data))
    except httpx.HTTPError as e:
        console.print(f"[yellow]Scooore skipped:[/yellow] {e}")
        return []


def fetch_napoleon_quotes(sport: str) -> list[OddQuote]:
    """Fetch + parse Napoleon's prematch 1X2 offer (Superbet platform, public
    REST, independent odds — genuinely widens value/surebet coverage)."""
    try:
        with NapoleonScraper() as nap:
            data = nap.fetch_by_date(sport)
        return list(napoleon_parse_by_date(data))
    except httpx.HTTPError as e:
        console.print(f"[yellow]Napoleon skipped:[/yellow] {e}")
        return []


def fetch_betfirst_quotes(sport: str) -> list[OddQuote]:
    """Fetch + parse the BetFirst events-table for a sport (paginated)."""
    try:
        with BetFirstScraper() as bf:
            data = bf.fetch_all_events(sport, days_ahead=7, max_market_count=10)
        return list(betfirst_parse_events_table(data))
    except httpx.HTTPError as e:
        console.print(f"[yellow]BetFirst skipped:[/yellow] {e}")
        return []


def fetch_ladbrokes_quotes(sport: str) -> list[OddQuote]:
    """Fetch + parse every Ladbrokes meeting of a sport via the detail-service."""
    try:
        with LadbrokesScraper() as lb:
            data = lb.fetch_all_meetings(sport, max_meetings=80)
        return list(ladbrokes_parse_prematch(data))
    except httpx.HTTPError as e:
        console.print(f"[yellow]Ladbrokes skipped:[/yellow] {e}")
        return []



def fetch_goldenpalace_quotes(sport: str) -> list[OddQuote]:
    """Bulk-fetch every Golden Palace event for a sport via the Altenar widget."""
    try:
        with GoldenPalaceScraper() as gp:
            data = gp.fetch_events(sport)
        return list(goldenpalace_parse_get_events(data))
    except httpx.HTTPError as e:
        console.print(f"[yellow]Golden Palace skipped:[/yellow] {e}")
        return []


def fetch_starcasinosport_quotes(sport: str) -> list[OddQuote]:
    """Bulk-fetch StarCasino Sport via the same Altenar widget."""
    try:
        with StarCasinoSportScraper() as ss:
            data = ss.fetch_events(sport)
        return list(starcasinosport_parse_get_events(data))
    except httpx.HTTPError as e:
        console.print(f"[yellow]StarCasino Sport skipped:[/yellow] {e}")
        return []


def fetch_betcenter_quotes(sport: str) -> list[OddQuote]:
    """Fetch + parse the Betcenter prematch offer (Cashpoint/Merkur platform,
    own odds API on oddsservice.betcenter.be -- independent odds)."""
    try:
        with BetCenterScraper() as bc:
            return bc.fetch_all_quotes(sport)
    except httpx.HTTPError as e:
        console.print(f"[yellow]Betcenter skipped:[/yellow] {e}")
        return []


def _fetch_all_parallel(
    sport: str,
    betano_file: str | None = None,
    *,
    include_file_books: bool = True,
) -> list[OddQuote]:
    """Fetch Pinnacle + all soft books concurrently. Returns the merged list;
    callers split by book to route Pinnacle/Smarkets as sharp references."""
    def _pinnacle() -> list[OddQuote]:
        with PinnacleScraper() as pin:
            return list(pin.fetch_market_quotes(sport))

    tasks: dict[str, Callable[[], list[OddQuote]]] = {
        "Pinnacle":      _pinnacle,
        "Unibet":        lambda: fetch_unibet_quotes(sport),
        # "711":           lambda: fetch_sevenelevenbe_quotes(sport),  # Kambi jumeau d'Unibet -> desactive (anti rate-limit)
        # "Bingoal":       lambda: fetch_bingoal_quotes(sport),  # Kambi jumeau d'Unibet -> desactive (anti rate-limit)
        # "Scooore":       lambda: fetch_scooore_quotes(sport),  # Kambi jumeau d'Unibet -> desactive (anti rate-limit)
        # MeridianBet: scraper prêt mais l'API exige un token (anti-bot
        # TrafficGuard) -> réactiver ici une fois le token capturé.
        # "MeridianBet": lambda: fetch_meridian_quotes(sport),
        # "BetFirst":      lambda: fetch_betfirst_quotes(sport),  # desactive temporairement
        "Ladbrokes":     lambda: fetch_ladbrokes_quotes(sport),
        "StarCasino":    lambda: fetch_starcasinosport_quotes(sport),
        "Napoleon":      lambda: fetch_napoleon_quotes(sport),
        # "Betcenter":     lambda: fetch_betcenter_quotes(sport),  # desactive: cotes erronees
        # Golden Palace retiré: compte limité, plus exploitable.
    }
    if include_file_books:
        tasks["Betano"]        = lambda: fetch_betano_quotes(betano_file=betano_file)

    all_quotes: list[OddQuote] = []
    with ThreadPoolExecutor(max_workers=len(tasks)) as executor:
        futures = {executor.submit(fn): name for name, fn in tasks.items()}
        for future in as_completed(futures):
            name = futures[future]
            try:
                quotes = future.result()
                all_quotes.extend(quotes)
                if quotes:
                    console.print(f"\\[{sport}]   → {len(quotes):5d} quotes  {name}")
            except Exception as e:
                console.print(f"[yellow]\\[{sport}]   {name} skipped: {e}[/yellow]")
    # Drop handicap quotes at the source: they're excluded from both value bets
    # and surebets, so parsing/storing/matching them is pure wasted CPU (and
    # they're ~45% of Pinnacle's payload). Filtering here lightens the whole
    # downstream pipeline — critical on a small CPU.
    return [q for q in all_quotes if q.market != MarketType.HANDICAP]


@app.command()
def scan(
    sport: str = "soccer",
    min_ev: float = 2.0,
    bankroll: float = 1000.0,
    betano_file: str = typer.Option(
        None,
        "--betano-file",
        help="Path to a JSON dump of Betano's /danae-webapi/.../live/overview/latest "
        "response, captured from your browser. Bypasses the IP-bound cookie check.",
    ),
):
    """Fetch Pinnacle + soft books (Betano, Unibet, 711, Bingoal, BetFirst, Ladbrokes, StarCasino), compute fair lines, print top value bets.

    --sport accepts a comma-separated list (e.g. 'soccer,tennis,basketball').
    The full pipeline runs per sport and results are tagged in their own
    section so per-sport coverage stays visible."""
    sports = [s.strip() for s in sport.split(",") if s.strip()]
    storage = Storage(ScanConfig().db_path)
    teams.init(storage)

    for current_sport in sports:
        cfg = ScanConfig(sport=current_sport, min_ev_pct=min_ev, bankroll=bankroll)
        console.print()
        console.print(f"[bold green]══ {current_sport.upper()} ══[/bold green]")

        console.print(f"[bold]Fetching all books in parallel ({current_sport})...[/bold]")
        all_quotes = _fetch_all_parallel(
            current_sport, betano_file,
            include_file_books=(current_sport == sports[0]),
        )

        quotes         = [q for q in all_quotes if q.book == Book.PINNACLE]
        raw_soft       = [q for q in all_quotes if q.book != Book.PINNACLE]

        fair = build_fair_lines(quotes, cfg.devig_method)
        console.print(f"  → {len(fair)} fair lines (devig={cfg.devig_method}, sharp=Pinnacle)")

        storage.insert_quotes(quotes)

        sport = current_sport  # keep local var name for downstream prints

        ref_keys = {fl.event_key for fl in fair.values()}
        soft_quotes = remap_to_reference(raw_soft, ref_keys)
        console.print(f"  → {len(soft_quotes)} matched to a Pinnacle event")
        storage.insert_quotes(soft_quotes)

        bets = merge_twin_book_value_bets(find_value_bets(soft_quotes, fair, cfg))
        bets.sort(key=lambda b: b.ev_pct, reverse=True)
        console.print(f"[bold]Value bets: {len(bets)}[/bold]")

        for b in bets:
            storage.insert_value_bet(b)

        tg_cfg = TelegramConfig.from_env()
        if tg_cfg is not None:
            candidates = [
                b for b in bets
                if storage.value_bet_notify_count(
                    b.event_key, b.book.value, b.market.value, b.outcome.label, b.outcome.line,
                ) < tg_cfg.valuebet_max_alerts
                and (
                    not tg_cfg.valuebet_dedup
                    or not storage.value_bet_already_notified(
                        b.event_key, b.book.value, b.market.value, b.outcome.label, b.outcome.line,
                        current_ev_pct=b.ev_pct,
                        ev_delta_pct=tg_cfg.valuebet_ev_delta_pct,
                    )
                )
            ]
            sent = send_alerts(candidates, tg_cfg, print_fn=lambda s: console.print(f"[yellow]{s}[/yellow]"), sport=current_sport)
            now = datetime.now(timezone.utc)
            for b in sent:
                storage.mark_value_bet_notified(
                    b.event_key, b.book.value, b.market.value, b.outcome.label, b.outcome.line,
                    b.ev_pct, now,
                )
            if sent:
                console.print(f"  → {len(sent)} Telegram alerts sent (EV ≥ {tg_cfg.min_ev_pct:.1f}%)")

        table = Table(title=f"Value bets ({sport}, min_ev={min_ev}%)", show_lines=False)
        table.add_column("event_key", overflow="fold")
        table.add_column("book")
        table.add_column("market")
        table.add_column("outcome")
        table.add_column("odd", justify="right")
        table.add_column("fair", justify="right")
        table.add_column("EV%", justify="right")
        table.add_column("stake%", justify="right")
        for b in bets[:25]:
            line = f" {b.outcome.line}" if b.outcome.line is not None else ""
            table.add_row(
                b.event_key, b.book.value, b.market.value, f"{b.outcome.label}{line}",
                f"{b.odd_taken:.2f}", f"{b.fair_odd:.2f}", f"{b.ev_pct:.2f}", f"{b.kelly_stake_pct:.2f}",
            )
        console.print(table)

        # Cross-book surebet detection. We include Pinnacle's own quotes in
        # the candidate pool — they're rarely the best odd per outcome (tight
        # margin) but on the occasional event where Pinnacle is the only
        # source for a side, or its odd happens to drift, the arb shows up
        # in the table. find_surebets already enforces a distinct-book-per-leg
        # rule, so Pinnacle-vs-Pinnacle "arbs" can't slip through.
        # Canonicalize across all books so events Pinnacle doesn't price still
        # yield surebets when two soft books cover both sides.
        surebets = find_surebets(canonicalize_for_surebets(quotes, raw_soft))
        plausible = [s for s in surebets if not s.suspicious]
        flagged = [s for s in surebets if s.suspicious]
        console.print(
            f"[bold]Surebets: {len(plausible)} plausible[/bold]"
            + (f" (+ {len(flagged)} flagged as suspicious — likely matching bugs)" if flagged else "")
        )

        # Telegram surebet alerts. The candidate pool depends on whether the
        # user opted into seeing suspicious ones; dedup is configurable too,
        # so a user who wants every scan to re-confirm can disable it.
        # Final per-margin filtering happens inside the alerter.
        if tg_cfg is not None and surebets:
            candidates = surebets if tg_cfg.include_suspicious_surebets else plausible
            candidates = [
                s for s in candidates
                if storage.surebet_notify_count(s.event_key, s.market.value, s.line)
                < tg_cfg.surebet_max_alerts
                and (
                    not tg_cfg.surebet_dedup
                    or not storage.surebet_already_notified(
                        s.event_key, s.market.value, s.line,
                        current_margin_pct=s.margin * 100,
                        roi_delta_pct=tg_cfg.surebet_roi_delta_pct,
                    )
                )
            ]
            if candidates:
                sent = send_surebet_alerts(
                    candidates, tg_cfg,
                    print_fn=lambda x: console.print(f"[yellow]{x}[/yellow]"),
                    sport=current_sport,
                )
                now = datetime.now(timezone.utc)
                for s in sent:
                    storage.mark_surebet_notified(
                        s.event_key, s.market.value, s.line, s.margin * 100, now,
                    )
                if sent:
                    console.print(f"  → {len(sent)} surebet alerts sent")
        if plausible:
            st = Table(title=f"Surebets ({sport})", show_lines=False)
            st.add_column("event_key", overflow="fold")
            st.add_column("market")
            st.add_column("line")
            st.add_column("legs", overflow="fold")
            st.add_column("margin%", justify="right")
            st.add_column("ROI%", justify="right")
            for s in plausible[:15]:
                legs_str = " | ".join(
                    f"{label}={odd:.2f} ({book.value})" for label, (odd, book) in s.legs.items()
                )
                st.add_row(
                    s.event_key,
                    s.market.value,
                    str(s.line) if s.line is not None else "-",
                    legs_str,
                    f"{s.margin * 100:.2f}",
                    f"{s.roi * 100:.2f}",
                )
            console.print(st)


@app.command(name="scan-surebets")
def scan_surebets(
    sport: str = "soccer",
    betano_file: str = typer.Option(
        None, "--betano-file",
        help="Optional Betano dump path — same as in `scan`.",
    ),
):
    """Surebet sweep including Pinnacle, designed to run every 5-15 min.
    Pinnacle quotes are fetched and used as the canonical event-key reference,
    then included in the surebet candidate pool — same as the full `scan`.
    Comma-separated --sport lets one cron entry cover every sport you care about."""
    sports = [s.strip() for s in sport.split(",") if s.strip()]
    storage = Storage(ScanConfig().db_path)
    teams.init(storage)
    tg_cfg = TelegramConfig.from_env()

    for current_sport in sports:
        console.print()
        console.print(f"[bold green]══ {current_sport.upper()} (surebets) ══[/bold green]")

        console.print(f"[bold]Fetching all books in parallel ({current_sport})...[/bold]")
        all_quotes = _fetch_all_parallel(
            current_sport, betano_file,
            include_file_books=(current_sport == sports[0]),
        )

        pinnacle_quotes = [q for q in all_quotes if q.book == Book.PINNACLE]
        soft_quotes     = [q for q in all_quotes if q.book != Book.PINNACLE]

        if not all_quotes:
            continue

        # Canonicalize across ALL books so surebets surface even on events
        # Pinnacle doesn't price — soft books anchor each other when Pinnacle is
        # absent. Pinnacle stays in the pool so Pinnacle-leg arbs are detected.
        normalised_quotes = canonicalize_for_surebets(pinnacle_quotes, soft_quotes)
        console.print(f"  → {len(normalised_quotes)} quotes matched to a common event")

        surebets = find_surebets(normalised_quotes)
        plausible = [s for s in surebets if not s.suspicious]
        flagged = [s for s in surebets if s.suspicious]
        console.print(
            f"[bold]Surebets: {len(plausible)} plausible[/bold]"
            + (f" (+ {len(flagged)} suspicious)" if flagged else "")
        )

        if tg_cfg is None or not surebets:
            continue

        candidates = surebets if tg_cfg.include_suspicious_surebets else plausible
        candidates = [
            s for s in candidates
            if storage.surebet_notify_count(s.event_key, s.market.value, s.line)
            < tg_cfg.surebet_max_alerts
            and (
                not tg_cfg.surebet_dedup
                or not storage.surebet_already_notified(
                    s.event_key, s.market.value, s.line,
                    current_margin_pct=s.margin * 100,
                    roi_delta_pct=tg_cfg.surebet_roi_delta_pct,
                )
            )
        ]
        if not candidates:
            continue
        sent = send_surebet_alerts(
            candidates, tg_cfg,
            print_fn=lambda x: console.print(f"[yellow]{x}[/yellow]"),
            sport=current_sport,
        )
        now = datetime.now(timezone.utc)
        for s in sent:
            storage.mark_surebet_notified(
                s.event_key, s.market.value, s.line, s.margin * 100, now,
            )
        if sent:
            console.print(f"  → {len(sent)} surebet alerts sent")


def _daemon_scan_sport(
    current_sport: str,
    storage: Storage,
    tg_cfg: "TelegramConfig | None",
    min_ev: float,
    bankroll: float,
    betano_file: "str | None",
) -> None:
    """Fetch, analyse and alert for one sport. Runs inside a ThreadPoolExecutor
    so all sports execute concurrently every cycle. SQLite WAL mode lets multiple
    threads read simultaneously; concurrent writes serialise via the 10 s timeout
    built into Storage._conn(), so no external locking is needed."""
    vb_to_mark: list[ValueBet] = []
    sb_to_mark: list[Surebet] = []
    mid_to_mark: list[Middle] = []
    now_mark = datetime.now(timezone.utc)

    try:
        console.print(f"\n[bold]{current_sport.upper()}[/bold]")
        # Betano is a soccer-only file-based scraper.
        all_q = _fetch_all_parallel(
            current_sport, betano_file,
            include_file_books=(current_sport == "soccer"),
        )
        pinnacle_q = [q for q in all_q if q.book == Book.PINNACLE]
        soft_raw   = [q for q in all_q if q.book != Book.PINNACLE]

        if not pinnacle_q:
            console.print(f"[yellow]\\[{current_sport}]   No Pinnacle quotes — skipping[/yellow]")
            return

        cfg = ScanConfig(sport=current_sport, min_ev_pct=min_ev, bankroll=bankroll)
        fair = build_fair_lines(pinnacle_q, cfg.devig_method)
        # Persist the event (with its sport) for every Pinnacle event in the
        # reference frame. Value bets are keyed onto these same event_keys, so
        # this lets clv-report break CLV down per sport instead of "unknown".
        event_rows = []
        for ek in {q.event_key for q in pinnacle_q}:
            parsed = parse_event_key(ek)
            if parsed is not None:
                start, home_norm, away_norm = parsed
                event_rows.append((ek, current_sport, "", home_norm, away_norm, start.isoformat()))
        storage.upsert_events(event_rows)
        storage.insert_quotes(pinnacle_q)

        ref_keys = {fl.event_key for fl in fair.values()}
        soft_q = remap_to_reference(soft_raw, ref_keys)
        storage.insert_quotes(soft_q)

        # ── CLV pre-kickoff alerts ────────────────────────────────────
        if tg_cfg is not None and tg_cfg.clv_window_minutes > 0:
            now_utc = datetime.now(timezone.utc)
            pin_idx: dict[tuple, float] = {}
            for _q in pinnacle_q:
                if "::" in _q.event_key:
                    _d = _q.event_key[:8]
                    _t = _q.event_key.split("::", 1)[1]
                    pin_idx[(_d, _t, _q.market.value, _q.outcome.label, _q.outcome.line)] = _q.decimal_odd

            clv_pending: list[tuple] = []
            for _bet in storage.open_value_bets():
                _parsed = parse_event_key(_bet["event_key"])
                if _parsed is None:
                    continue
                _kickoff, _, _ = _parsed
                _mins = (_kickoff - now_utc).total_seconds() / 60
                if not (0 < _mins <= tg_cfg.clv_window_minutes):
                    continue
                if storage.clv_alert_already_notified(int(_bet["id"])):
                    continue
                if "::" not in _bet["event_key"]:
                    continue
                _d = _bet["event_key"][:8]
                _t = _bet["event_key"].split("::", 1)[1]
                _pin_odd = pin_idx.get((_d, _t, _bet["market"], _bet["outcome_label"], _bet["line"]))
                if _pin_odd is None:
                    continue
                # Only alert on bet odds in the configured range (default
                # 1.5-4.0): below 1.5 the stake-to-reward is poor, above 4.0
                # the variance is too high for a small bankroll.
                _odd_taken = float(_bet["odd_taken"])
                if not (tg_cfg.clv_min_odd <= _odd_taken <= tg_cfg.clv_max_odd):
                    continue
                _clv = (_odd_taken / _pin_odd - 1) * 100
                if _clv < tg_cfg.min_clv_pct:
                    continue
                clv_pending.append((_bet, _clv, _pin_odd, int(_mins)))

            if clv_pending:
                clv_sent = send_clv_alerts(
                    clv_pending, tg_cfg,
                    print_fn=lambda s: console.print(f"[yellow]{s}[/yellow]"),
                    sport=current_sport,
                )
                # Mark only the CLV alerts that actually went out, so a
                # rate-limited/failed send is retried on a later cycle.
                for _bet_row, _clv_pct, _pin_odd, _mins in clv_sent:
                    storage.mark_clv_alert_notified(int(_bet_row["id"]), _clv_pct, _pin_odd, now_utc)
                if clv_sent:
                    console.print(f"\\[{current_sport}]   → {len(clv_sent)} CLV alert(s) sent")

        # ── Value bets ───────────────────────────────────────────────
        bets = merge_twin_book_value_bets(find_value_bets(soft_q, fair, cfg))
        bets.sort(key=lambda b: b.ev_pct, reverse=True)
        for b in bets:
            storage.insert_value_bet(b)
        console.print(f"\\[{current_sport}]   value bets: {len(bets)} total")
        if tg_cfg is not None:
            # The hard alert cap ALWAYS applies (even with the EV-delta dedup
            # turned off); the EV-delta dedup is an extra filter on top.
            vb_candidates = [
                b for b in bets
                if storage.value_bet_notify_count(
                    b.event_key, b.book.value, b.market.value, b.outcome.label, b.outcome.line,
                ) < tg_cfg.valuebet_max_alerts
                and (
                    not tg_cfg.valuebet_dedup
                    or not storage.value_bet_already_notified(
                        b.event_key, b.book.value, b.market.value, b.outcome.label, b.outcome.line,
                        current_ev_pct=b.ev_pct,
                        ev_delta_pct=tg_cfg.valuebet_ev_delta_pct,
                    )
                )
            ]
            sent = send_alerts(
                vb_candidates, tg_cfg,
                print_fn=lambda s: console.print(f"[yellow]{s}[/yellow]"),
                sport=current_sport,
            )
            now_mark = datetime.now(timezone.utc)
            # Mark only what actually sent — deferred/rate-limited bets stay
            # unmarked and get retried next cycle instead of being lost.
            vb_to_mark = sent
            if sent:
                console.print(f"\\[{current_sport}]   → {len(sent)} value bet alert(s) sent")

        # ── Surebets ─────────────────────────────────────────────────
        # Surebets use a wider pool than value bets: events Pinnacle doesn't
        # price still count, as long as two distinct books cover both sides.
        surebet_pool = canonicalize_for_surebets(pinnacle_q, soft_raw)
        surebets = find_surebets(surebet_pool)
        plausible = [s for s in surebets if not s.suspicious]
        console.print(f"\\[{current_sport}]   surebets: {len(plausible)} plausible")
        if tg_cfg is not None and surebets:
            sb_pool = surebets if tg_cfg.include_suspicious_surebets else plausible
            # Same dedup model as value bets: a hard alert cap that ALWAYS
            # applies, plus the optional ROI-delta dedup on top.
            sb_candidates = [
                s for s in sb_pool
                if storage.surebet_notify_count(s.event_key, s.market.value, s.line)
                < tg_cfg.surebet_max_alerts
                and (
                    not tg_cfg.surebet_dedup
                    or not storage.surebet_already_notified(
                        s.event_key, s.market.value, s.line,
                        current_margin_pct=s.margin * 100,
                        roi_delta_pct=tg_cfg.surebet_roi_delta_pct,
                    )
                )
            ]
            if sb_candidates:
                sent_sb = send_surebet_alerts(
                    sb_candidates, tg_cfg,
                    print_fn=lambda x: console.print(f"[yellow]{x}[/yellow]"),
                    sport=current_sport,
                )
                sb_to_mark = sent_sb
                if sent_sb:
                    console.print(f"\\[{current_sport}]   → {len(sent_sb)} surebet alert(s) sent")

        # ── Middles ──────────────────────────────────────────────────
        # Totals middles priced against Pinnacle's devigged ladder. Uses the
        # remapped soft quotes (aligned to the Pinnacle reference keys) so the
        # gap probability lookup hits the same event_key as `fair`.
        if tg_cfg is not None:
            middles = find_middles(
                soft_q, fair,
                min_ev_pct=tg_cfg.min_middle_ev_pct,
                max_gap=tg_cfg.middle_max_gap,
            )
            console.print(f"\\[{current_sport}]   middles: {len(middles)}")
            if middles:
                mid_candidates = [
                    m for m in middles
                    if storage.middle_notify_count(m.event_key, m.low_line, m.high_line)
                    < tg_cfg.middle_max_alerts
                    and (
                        not tg_cfg.middle_dedup
                        or not storage.middle_already_notified(
                            m.event_key, m.low_line, m.high_line,
                            current_ev_pct=m.ev_pct,
                            ev_delta_pct=tg_cfg.middle_ev_delta_pct,
                        )
                    )
                ]
                if mid_candidates:
                    sent_mid = send_middle_alerts(
                        mid_candidates, tg_cfg,
                        print_fn=lambda x: console.print(f"[yellow]{x}[/yellow]"),
                        sport=current_sport,
                    )
                    mid_to_mark = sent_mid
                    if sent_mid:
                        console.print(f"\\[{current_sport}]   → {len(sent_mid)} middle alert(s) sent")

    except Exception as e:
        console.print(f"[red]  {current_sport} error: {e}[/red]")

    # ── Persist dedup marks outside the sport catch so a scraper/analysis
    # failure never prevents already-sent alerts from being recorded. ──────
    try:
        for b in vb_to_mark:
            storage.mark_value_bet_notified(
                b.event_key, b.book.value, b.market.value, b.outcome.label, b.outcome.line,
                b.ev_pct, now_mark,
            )
        for s in sb_to_mark:
            storage.mark_surebet_notified(
                s.event_key, s.market.value, s.line, s.margin * 100, now_mark,
            )
        for m in mid_to_mark:
            storage.mark_middle_notified(
                m.event_key, m.low_line, m.high_line, m.ev_pct, now_mark,
            )
    except Exception as mark_err:
        console.print(
            f"[red]  dedup mark failed for {current_sport} — "
            f"next cycle may re-alert: {mark_err}[/red]"
        )


@app.command()
def daemon(
    sport: str = "soccer,tennis,basketball,hockey,volleyball",
    min_ev: float = typer.Option(
        5.0, "--min-ev",
        help="Minimum EV% to detect/store a value bet. Defaults to 5 to match "
        "the main chat's lower bound; the 5-8% bucket still shows ~+12% CLV. "
        "Below 5% is where the volume that re-saturated Telegram lives.",
    ),
    bankroll: float = 1000.0,
    breather: int = typer.Option(
        10, "--breather",
        help="Seconds to pause between cycle end and next start.",
    ),
    betano_file: str = typer.Option(
        None, "--betano-file",
        help="Path to a Betano JSON dump — same as in `scan`.",
    ),
):
    """Continuous scan: fetch all books in parallel, detect value bets + surebets,
    alert on Telegram only when something new or changed. Loops forever — run
    under systemd (scripts/valuebet-daemon.service) or screen/tmux.

    All sports are scanned concurrently (one thread per sport) so the cycle time
    equals the slowest single-sport fetch, not their sum."""
    sports_list = [s.strip() for s in sport.split(",") if s.strip()]
    storage = Storage(ScanConfig().db_path)
    teams.init(storage)

    cycle = 0
    while True:
        cycle += 1
        t0 = datetime.now(timezone.utc)
        console.print(f"\n[bold green]══ CYCLE {cycle} — {t0.strftime('%H:%M:%S')} UTC ══[/bold green]")
        tg_cfg = TelegramConfig.from_env()

        # Run every sport concurrently — cycle time = max(sport_time) not sum.
        with ThreadPoolExecutor(max_workers=len(sports_list)) as executor:
            futs = {
                executor.submit(
                    _daemon_scan_sport,
                    sp, storage, tg_cfg, min_ev, bankroll, betano_file,
                ): sp
                for sp in sports_list
            }
            for f in as_completed(futs):
                try:
                    f.result()
                except Exception as e:
                    console.print(f"[red]Sport thread {futs[f]} crashed: {e}[/red]")

        elapsed = (datetime.now(timezone.utc) - t0).total_seconds()
        console.print(f"\n[dim]Cycle {cycle} done in {elapsed:.0f}s — next in {breather}s[/dim]")
        time.sleep(breather)


@app.command(name="alert-test")
def alert_test():
    """Send one dummy alert per Telegram channel to verify all chats are wired up:
    value bet → TELEGRAM_CHAT_ID, surebet → TELEGRAM_SUREBET_CHAT_ID,
    CLV → TELEGRAM_CLV_CHAT_ID (each falls back to the main chat if not set)."""
    cfg = TelegramConfig.from_env()
    if cfg is None:
        console.print(
            "[yellow]TELEGRAM_BOT_TOKEN and/ou TELEGRAM_CHAT_ID non définis "
            "— rien à envoyer.[/yellow]"
        )
        return

    test_ev  = cfg.min_ev_pct + 1.0
    test_roi = cfg.min_surebet_margin_pct / 100 + 0.005

    sample_bet = ValueBet(
        event_key="202606010000::testteamA__vs__testteamB",
        book=Book.UNIBET_BE,
        market=MarketType.H2H,
        outcome=Outcome(label="home"),
        odd_taken=1.86,
        fair_prob=0.5650,
        fair_odd=1.77,
        ev_pct=test_ev,
        kelly_stake_pct=1.50,
        detected_at=datetime.now(timezone.utc),
    )
    # Prematch surebet: kickoff in the future
    sample_surebet_prematch = Surebet(
        event_key="202607010000::testteamA__vs__testteamB",
        market=MarketType.H2H,
        line=None,
        legs={
            "home": (1.95, Book.UNIBET_BE),
            "draw": (3.85, Book.BETFIRST),
            "away": (4.20, Book.LADBROKES_BE),
        },
        margin=test_roi,
    )
    # Live surebet: kickoff in the past
    sample_surebet_live = Surebet(
        event_key="202601010000::testteamA__vs__testteamB",
        market=MarketType.H2H,
        line=None,
        legs={
            "home": (1.95, Book.UNIBET_BE),
            "draw": (3.85, Book.BETFIRST),
            "away": (4.20, Book.LADBROKES_BE),
        },
        margin=test_roi,
    )
    # CLV test: dummy bet row as plain dict (same interface as sqlite3.Row)
    sample_clv_bet: dict = {
        "id": 0,
        "event_key": "202607010000::testteamA__vs__testteamB",
        "book": Book.UNIBET_BE.value,
        "market": MarketType.H2H.value,
        "outcome_label": "home",
        "line": None,
        "odd_taken": 1.86,
        "kelly_pct": 1.50,
    }
    # Premium channel samples: a big value bet (>min_premium_ev, odds in band)
    # and a juicy prematch surebet (>min_premium_surebet).
    premium_ev = max(cfg.min_premium_ev_pct + 1.0, cfg.min_ev_pct + 1.0)
    sample_premium_bet = ValueBet(
        event_key="202607010000::testteamA__vs__testteamB",
        book=Book.UNIBET_BE,
        market=MarketType.H2H,
        outcome=Outcome(label="home"),
        odd_taken=2.40,  # within the 1.5-4.0 premium band
        fair_prob=0.5650,
        fair_odd=1.77,
        ev_pct=premium_ev,
        kelly_stake_pct=1.50,
        detected_at=datetime.now(timezone.utc),
    )
    sample_premium_surebet = Surebet(
        event_key="202607010000::testteamA__vs__testteamB",
        market=MarketType.H2H,
        line=None,
        legs={
            "home": (1.95, Book.UNIBET_BE),
            "draw": (3.85, Book.BETFIRST),
            "away": (4.20, Book.LADBROKES_BE),
        },
        margin=cfg.min_premium_surebet_pct / 100 + 0.005,
    )

    bet_sent = send_alerts(
        [sample_bet], cfg, print_fn=lambda s: console.print(f"[yellow]{s}[/yellow]"),
        sport="soccer",
    )
    sb_prematch_sent = send_surebet_alerts(
        [sample_surebet_prematch], cfg,
        print_fn=lambda s: console.print(f"[yellow]{s}[/yellow]"),
        sport="soccer",
    )
    sb_live_sent = send_surebet_alerts(
        [sample_surebet_live], cfg,
        print_fn=lambda s: console.print(f"[yellow]{s}[/yellow]"),
        sport="soccer",
    )
    # CLV normal (7%) et CLV élevé (18%) — désormais le même canal, seul le
    # header change (🔥 au-dessus de min_high_clv_pct).
    clv_sent = send_clv_alerts(
        [(sample_clv_bet, 7.0, 1.74, 12)], cfg,
        print_fn=lambda s: console.print(f"[yellow]{s}[/yellow]"),
        sport="soccer",
    )
    clv_high_sent = send_clv_alerts(
        [(sample_clv_bet, 18.0, 1.58, 8)], cfg,
        print_fn=lambda s: console.print(f"[yellow]{s}[/yellow]"),
        sport="soccer",
    )
    # Premium : grosse value (envoyée aussi au canal principal) + surebet prématch.
    premium_bet_sent = send_alerts(
        [sample_premium_bet], cfg,
        print_fn=lambda s: console.print(f"[yellow]{s}[/yellow]"),
        sport="soccer",
    )
    premium_sb_sent = send_surebet_alerts(
        [sample_premium_surebet], cfg,
        print_fn=lambda s: console.print(f"[yellow]{s}[/yellow]"),
        sport="soccer",
    )

    def _status(sent: bool, chat: str, label: str, fallback_note: str = "") -> None:
        if sent:
            console.print(f"[bold]{label} → chat {chat} ✓{fallback_note}[/bold]")
        else:
            console.print(f"[red]{label} NOT sent — check the messages above.[/red]")

    _status(bet_sent, cfg.chat_id, "Value bet")

    sb_chat = cfg.effective_surebet_chat_id
    _status(
        sb_prematch_sent, sb_chat, "Surebet prématch",
        " (même chat — TELEGRAM_SUREBET_CHAT_ID non défini)" if sb_chat == cfg.chat_id else "",
    )

    live_chat = cfg.effective_live_surebet_chat_id
    _status(
        sb_live_sent, live_chat, "Surebet live",
        " (même chat — TELEGRAM_LIVE_SUREBET_CHAT_ID non défini)" if live_chat == sb_chat else "",
    )

    clv_chat = cfg.effective_clv_chat_id
    _status(
        clv_sent, clv_chat, "CLV normal (7%)",
        " (même chat — TELEGRAM_CLV_CHAT_ID non défini)" if clv_chat == cfg.chat_id else "",
    )
    _status(
        clv_high_sent, clv_chat, "CLV élevé (18%, header 🔥)",
        " (même canal CLV)",
    )

    premium_chat = cfg.effective_premium_chat_id
    if premium_chat:
        _status(premium_bet_sent, premium_chat, "Premium value (grosse EV)")
        _status(premium_sb_sent, premium_chat, "Premium surebet prématch")
    else:
        console.print(
            "[dim]Premium: TELEGRAM_PREMIUM_CHAT_ID non défini — canal premium désactivé.[/dim]"
        )


@app.command(name="close-lines")
def close_lines(sport: str = "soccer"):
    """For every detected value bet whose event has kicked off, snapshot the
    last Pinnacle price as the closing line. The closing price comes from our
    own historical capture in the quotes table — Pinnacle removes prematch
    markets from the live API at kickoff, so by the time this command runs
    the only place the real closing line still exists is in the rows scan
    persisted before kickoff. Run after kickoff (e.g. cron a few minutes
    past every hour); the closing snapshot feeds `clv-report`."""
    cfg = ScanConfig(sport=sport)
    storage = Storage(cfg.db_path)
    teams.init(storage)
    open_bets = storage.open_value_bets()
    if not open_bets:
        console.print("[bold]No open value bets to close.[/bold]")
        return

    now = datetime.now(timezone.utc)
    due = [b for b in open_bets if event_started(b["event_key"], now=now)]
    console.print(f"[bold]{len(open_bets)} open bets, {len(due)} past kickoff.[/bold]")
    if not due:
        return

    closed = 0
    missing = 0
    for b in due:
        parsed = parse_event_key(b["event_key"])
        if parsed is None:
            missing += 1
            continue
        kickoff, _, _ = parsed
        row = storage.latest_pinnacle_quote_before(
            event_key=b["event_key"],
            market=b["market"],
            outcome_label=b["outcome_label"],
            line=b["line"],
            before=kickoff,
        )
        if row is None:
            missing += 1
            continue
        pinnacle_odd = float(row["decimal_odd"])
        storage.insert_clv_snapshot(
            value_bet_id=int(b["id"]),
            pinnacle_odd=pinnacle_odd,
            # Inverse of the closing odd is the quickest implied-prob estimate
            # — close to the devigged fair but not normalised; the fair_prob
            # column on the value_bet row is the proper sharp estimate.
            pinnacle_prob=1.0 / pinnacle_odd,
            snapshot_at=now,
            closing=True,
        )
        closed += 1
    console.print(
        f"  → {closed} closing snapshots written; {missing} bets with no "
        f"pre-kickoff Pinnacle quote on file"
    )


_EV_BUCKET_ORDER = ["<5%", "5-8%", "8-15%", "15-35%", "35%+"]


def _ev_bucket(ev: float) -> str:
    """Bucket a detected EV% into the bands that map to the alert channels, so
    clv-report can show whether higher detected EV actually means better CLV."""
    if ev < 5:
        return "<5%"
    if ev < 8:
        return "5-8%"
    if ev < 15:
        return "8-15%"
    if ev < 35:
        return "15-35%"
    return "35%+"


@app.command()
def prune(
    retention_days: int = typer.Option(
        7, "--days",
        help="Delete raw quote rows older than this many days, then reclaim "
        "disk with VACUUM. 7 is safe: closing lines are captured within hours "
        "of kickoff. Run from cron to keep the DB from filling the disk.",
    ),
    vacuum: bool = typer.Option(True, "--vacuum/--no-vacuum",
                                help="Run VACUUM to actually shrink the file on disk."),
):
    """Trim the unbounded quotes history and reclaim disk space."""
    storage = Storage(ScanConfig().db_path)
    db_path = ScanConfig().db_path

    def _size_mb(p: str) -> float:
        try:
            return os.path.getsize(p) / (1024 * 1024)
        except OSError:
            return 0.0

    before = _size_mb(db_path)
    q = storage.prune_quotes(retention_days)
    n = storage.prune_notifications()
    console.print(f"Deleted {q} quote rows (> {retention_days}d) and {n} stale dedup rows.")
    if vacuum:
        console.print("VACUUM en cours (peut prendre un moment)…")
        storage.vacuum()
    after = _size_mb(db_path)
    console.print(f"DB : {before:.0f} Mo → {after:.0f} Mo (récupéré {before - after:.0f} Mo)")


@app.command(name="clv-report")
def clv_report():
    """Aggregate Closing Line Value over every closed value bet. CLV is the
    single most reliable indicator of long-run profitability — if your mean
    CLV is positive and stable, the engine is finding real edges."""
    cfg = ScanConfig()
    storage = Storage(cfg.db_path)
    teams.init(storage)
    rows = [dict(r) for r in storage.all_closed_bets()]
    if not rows:
        console.print("[bold]No closed bets yet — run `close-lines` after kickoffs.[/bold]")
        return

    pairs = [(r["odd_taken"], r["closing_odd"]) for r in rows]
    overall = clv_aggregate(pairs)
    console.print(
        f"[bold]Overall:[/bold] n={overall.n}  mean CLV {overall.mean_clv_pct:+.2f}%  "
        f"median {overall.median_clv_pct:+.2f}%  positive {overall.positive_rate * 100:.1f}%"
    )

    # Normalise the sport label (the LEFT JOIN yields None when the event row
    # was never persisted) and tag each row with its EV bucket.
    for r in rows:
        r["sport"] = r["sport"] or "unknown"
        r["ev_bucket"] = _ev_bucket(float(r["ev_pct"]))

    def _print_clv_table(title: str, dim: str, order: list[str] | None = None) -> None:
        groups = clv_group_by(rows, dim)
        stats = {k: clv_aggregate(v) for k, v in groups.items() if v}
        if not stats:
            return
        t = Table(title=title, show_lines=False)
        t.add_column(dim)
        t.add_column("n", justify="right")
        t.add_column("mean CLV%", justify="right")
        t.add_column("median%", justify="right")
        t.add_column("positive%", justify="right")
        # EV buckets read best in natural order; the rest by mean CLV descending.
        keys = ([k for k in order if k in stats] if order
                else sorted(stats, key=lambda k: stats[k].mean_clv_pct, reverse=True))
        for k in keys:
            s = stats[k]
            t.add_row(
                str(k), str(s.n),
                f"{s.mean_clv_pct:+.2f}", f"{s.median_clv_pct:+.2f}",
                f"{s.positive_rate * 100:.1f}",
            )
        console.print(t)

    _print_clv_table("CLV by book", "book")
    _print_clv_table("CLV by market", "market")
    _print_clv_table("CLV by sport", "sport")
    _print_clv_table("CLV by EV bucket", "ev_bucket", order=_EV_BUCKET_ORDER)


@app.command(name="export-history")
def export_history(
    out: str = typer.Option("history.csv", "--out", help="Chemin du fichier CSV de sortie."),
    bankroll: float = typer.Option(1000.0, "--bankroll", help="Capital de départ pour la simulation."),
):
    """Exporter chaque value bet clôturé (avec ligne de clôture + CLV) en CSV,
    prêt pour Excel : date, match, book, cotes, EV%, clôture, CLV%, mise, et un
    capital simulé sur la CLV. Colonnes Résultat/P&L laissées vides (à remplir
    à la main ou via un futur flux de résultats)."""
    import csv
    storage = Storage(ScanConfig().db_path)
    teams.init(storage)
    rows = [dict(r) for r in storage.all_closed_bets()]
    if not rows:
        console.print("[bold]Aucun pari clôturé — lance `close-lines` après des coups d'envoi.[/bold]")
        return

    # Oldest first so the simulated capital curve reads chronologically.
    rows.sort(key=lambda r: r.get("detected_at") or "")

    def _match(ek: str) -> str:
        parsed = parse_event_key(ek)
        if parsed is None:
            return ek
        _, home, away = parsed
        return f"{home.replace('_', ' ').title()} vs {away.replace('_', ' ').title()}"

    headers = [
        "Date", "Sport", "Match", "Book", "Marché", "Pari", "Cote prise",
        "Cote fair", "EV %", "Cote clôture (Pinnacle)", "CLV %", "Mise % (Kelly)",
        "Capital simulé (CLV)", "Résultat", "P&L réel",
    ]
    capital = bankroll
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for r in rows:
            taken = float(r["odd_taken"])
            closing = float(r["closing_odd"]) if r.get("closing_odd") else 0.0
            clv = clv_pct(taken, closing) * 100 if closing else 0.0
            kelly = float(r.get("kelly_pct") or 0.0)
            # Capital simulé : on mise kelly% du capital, gain espéré ≈ CLV.
            stake = capital * kelly / 100.0
            capital += stake * (clv / 100.0)
            line = r.get("line")
            pari = f"{r['outcome_label']}{(' ' + str(line)) if line is not None else ''}"
            w.writerow([
                (r.get("detected_at") or "")[:19],
                r.get("sport") or "",
                _match(r["event_key"]),
                r["book"],
                r["market"],
                pari,
                f"{taken:.2f}",
                f"{float(r['fair_odd']):.2f}",
                f"{float(r['ev_pct']):.2f}",
                f"{closing:.2f}" if closing else "",
                f"{clv:+.2f}" if closing else "",
                f"{kelly:.2f}",
                f"{capital:.2f}",
                "",   # Résultat (à remplir)
                "",   # P&L réel (à remplir)
            ])
    console.print(
        f"[green]✓[/green] {len(rows)} paris exportés vers [bold]{out}[/bold]  "
        f"(capital simulé CLV : {bankroll:.0f}€ → {capital:.0f}€)"
    )



@app.command(name="inspect-betano")
def inspect_betano(path: str):
    """Inspect a saved Betano overview JSON dump (DevTools → Response → save).

    Prints the shape of one event / league / market / selection so the parser
    can be refined for the exact field names that endpoint uses.
    """
    import json as _json
    from pathlib import Path as _Path

    data = _json.loads(_Path(path).read_text())

    def _peek(label: str, container: dict | None, n: int = 1) -> None:
        console.print(f"\n[bold cyan]{label}[/bold cyan] (count={len(container or {})})")
        if not container:
            return
        for i, (k, v) in enumerate(container.items()):
            console.print(f"  key={k!r}  fields={sorted((v or {}).keys()) if isinstance(v, dict) else type(v).__name__}")
            if isinstance(v, dict):
                console.print(f"  sample={_json.dumps(v, indent=2, ensure_ascii=False)[:600]}")
            if i + 1 >= n:
                break

    console.print(f"[bold]Top-level keys[/bold]: {sorted(data.keys())}")
    if "contentVersion" in data:
        console.print(f"contentVersion: {data['contentVersion']}")
    if "sports" in data and isinstance(data["sports"], dict):
        ids = data["sports"].get("allIds")
        if ids:
            console.print(f"sport ids: {ids}")

    _peek("events", data.get("events"))
    _peek("leagues", data.get("leagues"))
    _peek("markets", data.get("markets"), n=2)
    _peek("selections", data.get("selections"), n=2)
    _peek("zones", data.get("zones"))

    quotes = list(betano_parse_overview(data))
    console.print(f"\n[bold]Parser produced {len(quotes)} OddQuote(s).[/bold]")
    for q in quotes[:5]:
        console.print(f"  {q.event_key} | {q.market.value} | {q.outcome.label} @ {q.decimal_odd}")


@app.command(name="import-history")
def import_history(
    books: str = typer.Option(
        "", "--books",
        help="Books ciblés, séparés par des virgules (ex: unibet_be,bingoal_be). Vide = tous.",
    ),
    full: bool = typer.Option(
        False, "--full",
        help="Réimporte tout l'historique disponible au lieu de s'arrêter au dernier pari connu.",
    ),
    dump_raw: str = typer.Option(
        "", "--dump-raw",
        help="Sauvegarde le JSON brut du 1er book Kambi configuré dans ce fichier (pour décoder le schéma).",
    ),
):
    """Importer l'historique des paris réglés depuis l'API de chaque bookmaker
    vers la table settled_bets (suivi P&L réel). Idempotent : relançable, ne
    ré-insère jamais un pari déjà stocké. Voir `pnl-report` pour le bilan."""
    from datetime import datetime as _dt
    from .history.registry import importers_for
    from .history.base import HistoryImportError
    from .history.kambi import KambiHistoryImporter

    storage = Storage(ScanConfig().db_path)
    teams.init(storage)
    wanted = [b.strip() for b in books.split(",") if b.strip()]
    importers = importers_for(wanted)

    # --dump-raw : capturer le JSON brut pour verrouiller le schéma du parser.
    if dump_raw:
        kambi = next(
            (imp for imp in importers
             if isinstance(imp, KambiHistoryImporter) and imp.available()),
            None,
        )
        if kambi is None:
            console.print("[red]Aucun book Kambi configuré (URL + cookie) pour --dump-raw.[/red]")
            raise typer.Exit(1)
        try:
            data = kambi.raw_first_page()
        except HistoryImportError as e:
            console.print(f"[red]{e}[/red]")
            raise typer.Exit(1)
        import json as _json
        from pathlib import Path as _Path
        _Path(dump_raw).write_text(_json.dumps(data, indent=2, ensure_ascii=False))
        console.print(
            f"[green]✓[/green] JSON brut de [bold]{kambi.book.value}[/bold] écrit dans "
            f"[bold]{dump_raw}[/bold]. Top-level keys : "
            f"{sorted(data.keys()) if isinstance(data, dict) else type(data).__name__}"
        )
        return

    now = datetime.now(timezone.utc)
    total_new = 0
    for imp in importers:
        if not imp.available():
            note = getattr(imp, "note", "non configuré (URL/cookie manquant)")
            console.print(f"[dim]— {imp.book.value:<18} ignoré : {note}[/dim]")
            continue
        since = None
        if not full:
            last = storage.last_settled_at(imp.book.value)
            if last:
                try:
                    since = _dt.fromisoformat(last)
                except ValueError:
                    since = None
        try:
            bets = list(imp.fetch(since=since))
        except HistoryImportError as e:
            console.print(f"[red]✗ {imp.book.value:<18} {e}[/red]")
            continue
        except Exception as e:  # noqa: BLE001 — un book cassé ne doit pas tuer les autres
            console.print(f"[red]✗ {imp.book.value:<18} erreur inattendue : {e}[/red]")
            continue
        rows = [b.to_row(now) for b in bets]
        inserted, seen = storage.upsert_settled_bets(rows)
        total_new += inserted
        pnl = sum(b.pnl for b in bets)
        console.print(
            f"[green]✓[/green] {imp.book.value:<18} {inserted} nouveaux / {seen} vus "
            f"(P&L lot : {pnl:+.2f}€)"
        )
    console.print(f"\n[bold]{total_new} paris importés.[/bold] Bilan : [bold]valuebet pnl-report[/bold]")


@app.command(name="pnl-report")
def pnl_report():
    """Bilan P&L réel par book (depuis settled_bets), + total. Complète
    clv-report (qui mesure la CLV de ce qu'on a détecté) avec le vrai résultat
    encaissé sur chaque compte bookmaker."""
    storage = Storage(ScanConfig().db_path)
    rows = storage.pnl_by_book()
    if not rows:
        console.print(
            "[bold]Aucun pari importé.[/bold] Lance d'abord [bold]valuebet import-history[/bold] "
            "(après avoir renseigné les URL/cookies dans .env)."
        )
        return

    t = Table(title="P&L réel par book")
    for col in ("Book", "Paris", "En attente", "Misé (€)", "P&L (€)", "ROI %", "Win %"):
        t.add_column(col, justify="left" if col == "Book" else "right")

    tot_n = tot_pending = tot_wins = 0
    tot_staked = tot_pnl = 0.0
    for r in rows:
        n = int(r["n"] or 0)
        staked = float(r["staked"] or 0.0)
        pnl = float(r["pnl"] or 0.0)
        wins = int(r["wins"] or 0)
        pending = int(r["pending"] or 0)
        settled = n - pending
        roi = (pnl / staked * 100) if staked else 0.0
        win_pct = (wins / settled * 100) if settled else 0.0
        pnl_str = f"[green]{pnl:+.2f}[/green]" if pnl >= 0 else f"[red]{pnl:+.2f}[/red]"
        t.add_row(
            r["book"], str(n), str(pending), f"{staked:.2f}",
            pnl_str, f"{roi:+.1f}", f"{win_pct:.0f}",
        )
        tot_n += n
        tot_staked += staked
        tot_pnl += pnl
        tot_wins += wins
        tot_pending += pending

    tot_settled = tot_n - tot_pending
    tot_roi = (tot_pnl / tot_staked * 100) if tot_staked else 0.0
    tot_win = (tot_wins / tot_settled * 100) if tot_settled else 0.0
    tot_pnl_str = f"[green]{tot_pnl:+.2f}[/green]" if tot_pnl >= 0 else f"[red]{tot_pnl:+.2f}[/red]"
    t.add_section()
    t.add_row(
        "[bold]TOTAL[/bold]", f"[bold]{tot_n}[/bold]", str(tot_pending),
        f"[bold]{tot_staked:.2f}[/bold]", f"[bold]{tot_pnl_str}[/bold]",
        f"[bold]{tot_roi:+.1f}[/bold]", f"[bold]{tot_win:.0f}[/bold]",
    )
    console.print(t)


_RESULT_FR = {
    "won": "Gagné", "lost": "Perdu", "void": "Remboursé",
    "half_won": "Demi-gagné", "half_lost": "Demi-perdu",
    "cashout": "Cash-out", "pending": "En attente",
}


@app.command(name="pnl-export")
def pnl_export(
    out: str = typer.Option("paris.xlsx", "--out", help="Chemin du fichier Excel de sortie."),
):
    """Exporter tous les paris importés (table settled_bets) vers un fichier
    Excel : un onglet 'Paris' (une ligne par pari, du plus ancien au plus
    récent) et un onglet 'Bilan' (P&L par book + total). Régénéré à chaque
    appel — relance-le après chaque `import-history` pour un Excel à jour."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        console.print(
            "[red]openpyxl manquant.[/red] Installe-le : [bold]pip install openpyxl[/bold] "
            "(ou pip install -r requirements.txt)."
        )
        raise typer.Exit(1)

    storage = Storage(ScanConfig().db_path)
    rows = storage.settled_bets()  # newest first
    if not rows:
        console.print(
            "[bold]Aucun pari importé.[/bold] Lance d'abord [bold]valuebet import-history[/bold]."
        )
        raise typer.Exit(1)
    rows = list(reversed(rows))  # oldest first, so the sheet reads chronologically

    wb = Workbook()
    ws = wb.active
    ws.title = "Paris"
    headers = [
        "Date", "Bookmaker", "Sport", "Match", "Marché", "Sélection", "Cote",
        "Mise (€)", "Résultat", "Payout (€)", "P&L (€)", "Type", "Réglé le", "ID pari",
    ]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    win_fill = PatternFill("solid", fgColor="E2EFDA")   # vert clair
    loss_fill = PatternFill("solid", fgColor="FCE4E4")  # rouge clair

    def _short(ts):
        return (ts or "")[:16].replace("T", " ")

    for r in rows:
        legs = int(r["legs"] or 1)
        match = r["event_label"] or (f"Combiné {legs} sél." if legs > 1 else "—")
        sel = r["selection"] or ("—" if legs > 1 else "")
        line = r["line"]
        if line is not None and sel and sel != "—":
            lstr = f"{line:g}"          # 2.5 not 2.50
            if lstr not in sel:         # Kambi labels ("Plus de 2.5") already carry it
                sel = f"{sel} {lstr}"
        pnl = float(r["pnl"] or 0.0)
        ws.append([
            _short(r["placed_at"]),
            r["book"],
            (r["sport"] or "").title() if r["sport"] else "",
            match,
            r["market"] or ("Combiné" if legs > 1 else ""),
            sel,
            round(float(r["odd"] or 0.0), 3),
            round(float(r["stake"] or 0.0), 2),
            _RESULT_FR.get(r["result"], r["result"]),
            round(float(r["payout"]), 2) if r["payout"] is not None else None,
            round(pnl, 2),
            r["bet_type"] or "",
            _short(r["settled_at"]),
            r["bet_id"],
        ])
        row_idx = ws.max_row
        if pnl > 0:
            fill = win_fill
        elif pnl < 0:
            fill = loss_fill
        else:
            fill = None
        if fill:
            for cell in ws[row_idx]:
                cell.fill = fill

    # Largeurs de colonnes lisibles.
    widths = [17, 16, 12, 32, 20, 22, 8, 10, 12, 11, 10, 10, 17, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Onglet Bilan.
    bilan = wb.create_sheet("Bilan")
    bilan.append(["Bookmaker", "Paris", "En attente", "Misé (€)", "P&L (€)", "ROI %", "Win %"])
    for cell in bilan[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    tot_n = tot_pending = tot_wins = 0
    tot_staked = tot_pnl = 0.0
    for r in storage.pnl_by_book():
        n = int(r["n"] or 0)
        staked = float(r["staked"] or 0.0)
        pnl = float(r["pnl"] or 0.0)
        wins = int(r["wins"] or 0)
        pending = int(r["pending"] or 0)
        settled = n - pending
        bilan.append([
            r["book"], n, pending, round(staked, 2), round(pnl, 2),
            round(pnl / staked * 100, 1) if staked else 0.0,
            round(wins / settled * 100, 0) if settled else 0.0,
        ])
        tot_n += n; tot_pending += pending; tot_wins += wins
        tot_staked += staked; tot_pnl += pnl
    tot_settled = tot_n - tot_pending
    bilan.append([
        "TOTAL", tot_n, tot_pending, round(tot_staked, 2), round(tot_pnl, 2),
        round(tot_pnl / tot_staked * 100, 1) if tot_staked else 0.0,
        round(tot_wins / tot_settled * 100, 0) if tot_settled else 0.0,
    ])
    for cell in bilan[bilan.max_row]:
        cell.font = Font(bold=True)
    for i, w in enumerate([18, 8, 11, 10, 10, 8, 8], start=1):
        bilan.column_dimensions[get_column_letter(i)].width = w

    wb.save(out)
    console.print(
        f"[green]✓[/green] {len(rows)} paris exportés vers [bold]{out}[/bold] "
        f"(P&L total : {tot_pnl:+.2f}€). Ouvre-le dans Excel."
    )


@app.command(name="inspect-kambi-history")
def inspect_kambi_history(path: str):
    """Inspecter un dump JSON d'historique Kambi (capturé via --dump-raw ou
    DevTools → Response → save). Affiche la structure et ce que le parser en
    tire, pour verrouiller le mapping des champs."""
    import json as _json
    from pathlib import Path as _Path
    from .history.kambi import parse_history, _coupon_list, _outcomes

    data = _json.loads(_Path(path).read_text())
    console.print(
        f"[bold]Top-level keys[/bold]: "
        f"{sorted(data.keys()) if isinstance(data, dict) else type(data).__name__}"
    )
    coupons = _coupon_list(data)
    console.print(f"[bold]Coupons trouvés[/bold]: {len(coupons)}")
    if coupons:
        first = coupons[0]
        coupon = (
            first.get("coupon")
            if isinstance(first, dict) and isinstance(first.get("coupon"), dict)
            else first
        )
        console.print(
            f"[bold cyan]Champs coupon[/bold cyan]: "
            f"{sorted(coupon.keys()) if isinstance(coupon, dict) else '?'}"
        )
        console.print(_json.dumps(coupon, indent=2, ensure_ascii=False)[:1200])
        legs = _outcomes(coupon) if isinstance(coupon, dict) else []
        if legs and isinstance(legs[0], dict):
            console.print(f"\n[bold cyan]Champs 1ère sélection[/bold cyan]: {sorted(legs[0].keys())}")
            console.print(_json.dumps(legs[0], indent=2, ensure_ascii=False)[:800])

    bets = parse_history(data, Book.UNIBET_BE)
    console.print(f"\n[bold]Parser → {len(bets)} SettledBet.[/bold]")
    for b in bets[:5]:
        console.print(
            f"  {b.bet_id} | {b.result} | cote {b.odd} | mise {b.stake}€ | "
            f"P&L {b.pnl:+.2f}€ | {b.event_label or '—'}"
        )


@app.command()
def selftest():
    """Sanity check on math primitives."""
    from .devig import devig as _devig
    probs = _devig([2.10, 3.40, 3.80], method="shin")
    console.print(f"Shin devig of (2.10, 3.40, 3.80) → {probs}, sum={sum(probs):.6f}")
    console.print(f"EV at 2.10 vs fair_prob {probs[0]:.4f} = {ev_pct(2.10, probs[0]):.3f}%")
    console.print(f"Kelly fraction = {kelly_fraction(2.10, probs[0]):.4f}")
    console.print(f"Quarter Kelly stake on €1000 bankroll = €{kelly_stake(2.10, probs[0], 1000.0):.2f}")


if __name__ == "__main__":
    app()
