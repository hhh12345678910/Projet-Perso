from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any

EXCEL_PATH = Path("data/paris_joues.xlsx")

_HEADERS = [
    "Date joué", "Sport", "Match", "Coup d'envoi",
    "Book", "Marché", "Pari", "Cote",
    "EV%", "CLV%", "Marge%",
]


def _parse_kickoff(event_key: str) -> str:
    try:
        dt = datetime.strptime(event_key[:12], "%Y%m%d%H%M")
        return dt.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return ""


def _parse_matchup(event_key: str) -> str:
    try:
        teams_part = event_key.split("::", 1)[1]
        home, away = teams_part.split("__vs__", 1)
        return f"{home.replace('_', ' ').title()} vs {away.replace('_', ' ').title()}"
    except Exception:
        return event_key


def append_bet(alert: Any, played_at: datetime) -> None:
    """Write one row to the Excel tracker. Creates the file + header if needed."""
    import openpyxl
    from openpyxl.styles import Font

    EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)

    if EXCEL_PATH.exists():
        wb = openpyxl.load_workbook(str(EXCEL_PATH))
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Paris joués"
        ws.append(_HEADERS)
        for col_idx in range(1, len(_HEADERS) + 1):
            ws.cell(row=1, column=col_idx).font = Font(bold=True)

    alert_type = alert["alert_type"]
    event_key = alert["event_key"]
    matchup = _parse_matchup(event_key)
    kickoff = _parse_kickoff(event_key)
    sport = alert["sport"] or ""

    if alert_type == "surebet":
        legs_raw = json.loads(alert["legs_json"]) if alert["legs_json"] else {}
        legs_str = " / ".join(
            f"{label}={float(odd):.2f}({book})" for label, (odd, book) in legs_raw.items()
        )
        row = [
            played_at.strftime("%d/%m/%Y %H:%M"),
            sport,
            matchup,
            kickoff,
            "",
            alert["market"],
            legs_str,
            None,
            None,
            None,
            alert["margin_pct"],
        ]
    else:
        line_suffix = f" {alert['line']}" if alert["line"] is not None else ""
        outcome = alert["outcome_label"] or ""
        pari = f"{outcome}{line_suffix}".strip()
        row = [
            played_at.strftime("%d/%m/%Y %H:%M"),
            sport,
            matchup,
            kickoff,
            alert["book"] or "",
            alert["market"],
            pari,
            alert["odd_taken"],
            alert["ev_pct"],
            alert["clv_pct"],
            None,
        ]

    ws.append(row)
    wb.save(str(EXCEL_PATH))
